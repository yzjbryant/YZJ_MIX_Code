import os
import xlrd
import pandas as pd
from xlutils.copy import copy
import openpyxl
from openpyxl import load_workbook




# os.chdir('%s'%Folder_name[0])

path="F:\Ds_analyze"

Folder_name=[]
with open(os.path.join(path,'data.mdp'),'r') as f:
    for line in f.readlines()[0:1]:
        x=line[0:3]
        Folder_name.append(x)
print(Folder_name)



a=[]
b=[]
c=[]
d=[]
ds=[]
ds_error=[]
Gas_number=[]

with open(os.path.join(path,'data.mdp'),'r') as f:
    for line in f.readlines()[1:5]:
        x=line[0:8]
        ds.append(x)
        y=line[12:19]
        ds_error.append(y)
with open(os.path.join(path, 'data.mdp'), 'r') as f:
    for line in f.readlines()[5:206]:
        x=line[11:24]
        a.append(x)
with open(os.path.join(path, 'data.mdp'), 'r') as f:
    for line in f.readlines()[206:407]:
        x=line[11:24]
        b.append(x)
with open(os.path.join(path, 'data.mdp'), 'r') as f:
    for line in f.readlines()[407:608]:
        x=line[11:24]
        c.append(x)
with open(os.path.join(path, 'data.mdp'), 'r') as f:
    for line in f.readlines()[608:809]:
        x=line[11:24]
        d.append(x)
with open(os.path.join(path,'data.mdp'),'r') as f:
    for line in f.readlines()[809:813]:
        x=line[0:5]
        Gas_number.append(x)


# 进第一个sheet
workbook=xlrd.open_workbook(r'Ds_CH4.xlsx')
sheet=workbook.sheet_by_index(0)
row_count=sheet.nrows
col_count=sheet.ncols
# print(row_count)
# print(col_count)
rows=sheet.row_values(1)

Material_name=[]
for i in rows:
    if i !='':
        Material_name.append(i)
# print(Material_name)

col_single=[]
for i in Material_name:
    if i=='%s'%Folder_name[0]:
        x=(Material_name.index(i)-1)*5+3
        col_single.append(x)
        print(col_single)

wb = load_workbook("Ds_CH4.xlsx")
wb1 = wb.active

for i in range(len(ds)):
    wb1.cell(i+6, col_single[0], ds[i])
    wb1.cell(i+6, col_single[0]+1, ds_error[i])

for i in range(len(a)):  # 0-200
    wb1.cell(i+12, col_single[0], a[i])
    wb1.cell(i+12, col_single[0]+1, b[i])
    wb1.cell(i+12, col_single[0]+2, c[i])
    wb1.cell(i+12, col_single[0]+3, d[i])

for i in range(len(Gas_number)):
    wb1.cell(i+6,col_single[0]-1,Gas_number[i])

wb.save("Ds_CH4.xlsx")

my_file='F:\Ds_analyze\data.mdp'
if os.path.exists(my_file):
    os.remove(my_file)
else:
    print("No file")


# #写新文件
# wb=openpyxl.load_workbook(r'1.xlsx')
# ws=wb['Material']
# for i in range(1,len(a)+1):
#     data=a[i-1]
#     ws.cell(row=i+1,column=11).value=data
# wb.save("1.xlsx")

#获取sheet个数
# nums=len(workbook.sheets())
#获取sheet名称
# sheet_name=workbook.sheet_names()[0]

# demo_df=pd.read_excel(r'Ds_CO2.xlsx')
# for indexs in demo_df.index:
#     for i in range(len(demo_df.loc[indexs].values)):
#         if (demo_df.loc[indexs].values[i]=='%s'%Folder_name[0]):
#             row_single.append(indexs+2)
#             col_single.append(i+1)
#             print('行数：',indexs+2, '列数：',i+1)
#             print(demo_df.loc[indexs].values[i])
# demo_df.close()

# filename='Ds_CO2.xlsx'
# rb=xlrd.open_workbook(filename)
# wb=copy(rb)
# sheet=wb.get_sheet(0)
#
# for i in range(len(ds)):
#     sheet.write(i+5,col_single[0],ds[i])
#     sheet.write(i+5,col_single[0]+1,ds_error[i])
#
# for i in range(len(a)):#0-200
#     sheet.write(i+11,col_single[0],a[i])
#     sheet.write(i+11,col_single[0]+1,b[i])
#     sheet.write(i+11,col_single[0]+2,c[i])
#     sheet.write(i+11,col_single[0]+3,d[i])
#
# os.remove(filename)
# wb.save(filename)





