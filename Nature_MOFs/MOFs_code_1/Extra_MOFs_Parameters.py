import pymatgen as mg
import os
import xlrd
from openpyxl import load_workbook

####################################################################获取文件夹下CIF文件名
path='F:\MOFS_Extra_Data\cif\MOFs'
path2='F:\MOFS_Extra_Data'

path_list=os.listdir(path)
path_list.sort()
Cif_name=[]
for filename in path_list:
    Cif_name.append(filename)

####################################################################遍历CIF文件
Length_a=[]
Length_b=[]
Length_c=[]
Angle_alpha=[]
Angle_beta=[]
Angle_gamma=[]

for i in range(len(Cif_name)):
    with open(os.path.join(path,'%s'%Cif_name[i]),'r') as f1:
        for line in f1.readlines()[9:10]:
            x=line[30:45]
            Length_a.append(x)
    with open(os.path.join(path, '%s' % Cif_name[i]), 'r') as f1:
        for line in f1.readlines()[10:11]:
            x=line[30:45]
            Length_b.append(x)
    with open(os.path.join(path, '%s' % Cif_name[i]), 'r') as f1:
        for line in f1.readlines()[11:12]:
            x=line[30:45]
            Length_c.append(x)
    with open(os.path.join(path, '%s' % Cif_name[i]), 'r') as f1:
        for line in f1.readlines()[12:13]:
            x=line[30:45]
            Angle_alpha.append(x)
    with open(os.path.join(path, '%s' % Cif_name[i]), 'r') as f1:
        for line in f1.readlines()[13:14]:
            x=line[30:45]
            Angle_beta.append(x)
    with open(os.path.join(path, '%s' % Cif_name[i]), 'r') as f1:
        for line in f1.readlines()[14:15]:
            x=line[30:45]
            Angle_gamma.append(x)

#################################################################读pymatgen分析出来的数据
Element=['Tong','Zn','C','H','Br','Lu','I','N','O','F','S'] #Cu-Tong,Cl-Lu
Element2=[] #Number
Tong=[]
Zn=[]
C=[]
H=[]
Br=[]
Lu=[]
I=[]
N=[]
O=[]
F=[]
S=[]

with open(os.path.join(path2,'All_Element.txt'),'r') as f:
    Element3 = ['0' for index in range(11)]  # Number + 0 + ... +0
    for line in f.readlines():
        temp=line.split() #Number
        temp1=line.split() #Element+Number
        for j in range(len(temp)):
            for i in range(len(Element)):
                temp[j] = temp[j].lstrip(Element[i])
            Element2.append(temp[j])
        for j in range(len(temp)):
            for i in range(len(Element)):
                if Element[i] in temp1[j]:
                    Element3[i]=temp[j]
        # print(Element3)
        Tong.append('%s\n' % Element3[0])
        Zn.append('%s\n' % Element3[1])
        C.append('%s\n' % Element3[2])
        H.append('%s\n' % Element3[3])
        Br.append('%s\n' % Element3[4])
        Lu.append('%s\n' % Element3[5])
        I.append('%s\n' % Element3[6])
        N.append('%s\n' % Element3[7])
        O.append('%s\n' % Element3[8])
        F.append('%s\n' % Element3[9])
        S.append('%s\n' % Element3[10])
        with open(os.path.join(path2, 'Element_Number.txt'), 'a') as f:
            f.write('%s %s %s %s %s %s %s %s %s %s %s\n'%(Element3[0],Element3[1],Element3[2],Element3[3],
                                                    Element3[4], Element3[5],Element3[6],Element3[7],
                                                    Element3[8], Element3[9],Element3[10]))
    Element2=[]
    Element3=['0' for index in range(11)]

#################################################################写文件
wb=load_workbook("CIF_Parameters.xlsx")
wb1=wb.active

for i in range(len(Length_a)):
    wb1.cell(1,1,'Materials')
    wb1.cell(1,2,'Length_a')
    wb1.cell(1,3,'Length_b')
    wb1.cell(1,4,'Length_c')
    wb1.cell(1,5,'Angle_alpha')
    wb1.cell(1,6,'Angle_beta')
    wb1.cell(1,7,'Angle_gamma')
    wb1.cell(1,8,'Tong')
    wb1.cell(1,9,'Zn')
    wb1.cell(1,10,'C')
    wb1.cell(1,11,'H')
    wb1.cell(1,12,'Br')
    wb1.cell(1,13,'Lu')
    wb1.cell(1,14,'I')
    wb1.cell(1,15,'N')
    wb1.cell(1,16,'O')
    wb1.cell(1,17,'F')
    wb1.cell(1,18,'S')
    wb1.cell(1, 19, 'Energy_C4H4S_298K_10kpa')
    wb1.cell(1, 20, 'Energy_C4H4S_363K_10kpa')
    wb1.cell(1, 21, 'Energy_C4H4S_363K_100kpa')
    wb1.cell(1, 22, 'Energy_C6H6_298K_10kpa')
    wb1.cell(1, 23, 'Energy_C6H6_363K_10kpa')
    wb1.cell(1, 24, 'Energy_C6H6_363K_100kpa')
    wb1.cell(1, 25, 'Loading_C4H4S_298K_10kpa')
    wb1.cell(1, 26, 'Loading_C4H4S_363K_10kpa')
    wb1.cell(1, 27, 'Loading_C4H4S_363K_100kpa')
    wb1.cell(1, 28, 'Loading_C6H6_298K_10kpa')
    wb1.cell(1, 29, 'Loading_C6H6_363K_10kpa')
    wb1.cell(1, 30, 'Loading_C6H6_363K_100kpa')

    wb1.cell(i+2,1,Cif_name[i])
    wb1.cell(i+2,2,Length_a[i])
    wb1.cell(i+2,3,Length_b[i])
    wb1.cell(i+2,4,Length_c[i])
    wb1.cell(i+2,5,Angle_alpha[i])
    wb1.cell(i+2,6,Angle_beta[i])
    wb1.cell(i+2,7,Angle_gamma[i])
    wb1.cell(i + 2, 8, Tong[i])
    wb1.cell(i + 2, 9, Zn[i])
    wb1.cell(i + 2, 10, C[i])
    wb1.cell(i + 2, 11, H[i])
    wb1.cell(i + 2, 12, Br[i])
    wb1.cell(i + 2, 13, Lu[i])
    wb1.cell(i + 2, 14, I[i])
    wb1.cell(i + 2, 15, N[i])
    wb1.cell(i + 2, 16, O[i])
    wb1.cell(i + 2, 17, F[i])
    wb1.cell(i + 2, 18, S[i])
wb.save("CIF_Parameters.xlsx")

my_file='F:\MOFS_Extra_Data\Element_Number.txt'
if os.path.exists(my_file):
    os.remove(my_file)
else:
    print("No file")
############################################################################所有原子个数、种类
'''
Full_Formula=[]
for i in range(len(Cif_name)):
    structure=mg.Structure.from_file(os.path.join(path,'%s'%Cif_name[i]))
    doc=open('out.txt','w')
    print(structure,file=doc)
    with open(os.path.join(path2,'out.txt'),'r') as f:
        for line in f.readlines()[0:1]:
            x=line
            Full_Formula.append(x)
        f.close()
    print(Full_Formula)
with open(os.path.join(path2,'All.txt'),'w') as f:
    for i in range(len(Full_Formula)):
        f.write(Full_Formula[i])
'''

