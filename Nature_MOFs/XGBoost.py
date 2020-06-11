import numpy as np
import pandas as pd
import os
import sys
from pandas import DataFrame
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier,RandomForestRegressor,AdaBoostClassifier,AdaBoostRegressor
from sklearn.linear_model import LogisticRegression
from sklearn.svm import SVC,LinearSVC,LinearSVR,SVR,NuSVC,NuSVR,OneClassSVM
from sklearn.neighbors import KNeighborsClassifier,KNeighborsRegressor
from sklearn.naive_bayes import GaussianNB,MultinomialNB,BernoulliNB
from sklearn.ensemble import GradientBoostingClassifier,GradientBoostingRegressor
from sklearn.tree import DecisionTreeRegressor,DecisionTreeClassifier
from sklearn import preprocessing
import tensorflow as tf
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import Normalizer
from sklearn.linear_model import Ridge,Lasso,ElasticNet,LinearRegression
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
import json
import xgboost as xgb
from xgboost import XGBRegressor
from math import sqrt
#分类指标
from sklearn.metrics import accuracy_score,classification_report,confusion_matrix
# 回归指标
from sklearn.metrics import mean_absolute_error,mean_squared_error,r2_score
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import PolynomialFeatures,normalize
from sklearn.model_selection import KFold,cross_val_score

data_1=pd.read_excel('CIF_Parameters.xlsx',header=0)
'''
Feature_Descriptor=data.columns
print(Feature_Descriptor)
'''
Target=['Pure_C6H6_Heat_298K_9.99kpa',
       'Pure_C6H6_Heat_363K_9.99kpa', 'Pure_C6H6_Heat_363K_99.9kpa',
       'Pure_C6H6_Loading_298K_9.99kpa', 'Pure_C6H6_Loading_363K_9.99kpa',
       'Pure_C6H6_Loading_363K_99.9kpa', 'Pure_C4H4S_Heat_298K_0.01kpa',
       'Pure_C4H4S_Heat_363K_0.01kpa', 'Pure_C4H4S_Heat_363K_0.1kpa',
       'Pure_C4H4S_Loading_298K_0.01kpa', 'Pure_C4H4S_Loading_363K_0.01kpa',
       'Pure_C4H4S_Loading_363K_0.1kpa', 'Mix_Heat_C4H4S_298K_10kpa',
       'Mix_Heat_C4H4S_363K_10kpa', 'Mix_Heat_C4H4S_363K_100kpa',
       'Mix_Heat_C6H6_298K_10kpa', 'Mix_Heat_C6H6_363K_10kpa',
       'Mix_Heat_C6H6_363K_100kpa', 'Mix_Loading_C4H4S_298K_10kpa',
       'Mix_Loading_C4H4S_363K_10kpa', 'Mix_Loading_C4H4S_363K_100kpa',
       'Mix_Loading_C6H6_298K_10kpa', 'Mix_Loading_C6H6_363K_10kpa',
       'Mix_Loading_C6H6_363K_100kpa', 'Selectivity_298K_10kpa',
       'Selectivity_363K_10kpa', 'Selectivity_298K_100kpa',
       'Working_capacity_PSA_363K_100kpa_10kpa',
       'Working_capacity_TSA_10kpa_298K_363K']#Number of Target(i) is 29
Feature_Descriptor=['Cluster', 'Linker', 'Group', 'Mass', 'Length_a',
       'Length_b', 'Length_c', 'Angle_alpha', 'Angle_beta', 'Angle_gamma',
       'Number of Element', 'Tong', 'Zn', 'C', 'H', 'Br', 'Lu', 'I', 'N', 'O',
       'F', 'S', 'Di', 'Df', 'Dif', 'Unitcell_volume', 'Density', 'ASA_A^2',
       'ASA_m^2/cm^3', 'ASA_m^2/g', 'NASA_A^2', 'NASA_m^2/cm^3', 'NASA_m^2/g',
       'AV_A^3', 'AV_Volume_fraction', 'AV_cm^3/g', 'NAV_A^3',
       'NAV_Volume_fraction', 'NAV_cm^3/g', 'Pure_C6H6_Heat_298K_9.99kpa',
       'Pure_C6H6_Heat_363K_9.99kpa', 'Pure_C6H6_Heat_363K_99.9kpa',
       'Pure_C6H6_Loading_298K_9.99kpa', 'Pure_C6H6_Loading_363K_9.99kpa',
       'Pure_C6H6_Loading_363K_99.9kpa', 'Pure_C4H4S_Heat_298K_0.01kpa',
       'Pure_C4H4S_Heat_363K_0.01kpa', 'Pure_C4H4S_Heat_363K_0.1kpa',
       'Pure_C4H4S_Loading_298K_0.01kpa', 'Pure_C4H4S_Loading_363K_0.01kpa',
       'Pure_C4H4S_Loading_363K_0.1kpa', 'Mix_Heat_C4H4S_298K_10kpa',
       'Mix_Heat_C4H4S_363K_10kpa', 'Mix_Heat_C4H4S_363K_100kpa',
       'Mix_Heat_C6H6_298K_10kpa', 'Mix_Heat_C6H6_363K_10kpa',
       'Mix_Heat_C6H6_363K_100kpa', 'Mix_Loading_C4H4S_298K_10kpa',
       'Mix_Loading_C4H4S_363K_10kpa', 'Mix_Loading_C4H4S_363K_100kpa',
       'Mix_Loading_C6H6_298K_10kpa', 'Mix_Loading_C6H6_363K_10kpa',
       'Mix_Loading_C6H6_363K_100kpa', 'Selectivity_298K_10kpa',
       'Selectivity_363K_10kpa', 'Selectivity_298K_100kpa',
       'Working_capacity_PSA_363K_100kpa_10kpa',
       'Working_capacity_TSA_10kpa_298K_363K'] #Number of Feature_Descriptor is 69 'Materials',
# data=data.drop(['Materials'],axis=1)
data=data_1.drop(['Materials','Pure_C6H6_Heat_298K_9.99kpa',
       'Pure_C6H6_Heat_363K_9.99kpa', 'Pure_C6H6_Heat_363K_99.9kpa',
       'Pure_C6H6_Loading_298K_9.99kpa', 'Pure_C6H6_Loading_363K_9.99kpa',
       'Pure_C6H6_Loading_363K_99.9kpa', 'Pure_C4H4S_Heat_298K_0.01kpa',
       'Pure_C4H4S_Heat_363K_0.01kpa', 'Pure_C4H4S_Heat_363K_0.1kpa',
       'Pure_C4H4S_Loading_298K_0.01kpa', 'Pure_C4H4S_Loading_363K_0.01kpa',
       'Pure_C4H4S_Loading_363K_0.1kpa', 'Mix_Heat_C4H4S_298K_10kpa',
       'Mix_Heat_C4H4S_363K_10kpa', 'Mix_Heat_C4H4S_363K_100kpa',
       'Mix_Heat_C6H6_298K_10kpa', 'Mix_Heat_C6H6_363K_10kpa',
       'Mix_Heat_C6H6_363K_100kpa', 'Mix_Loading_C4H4S_298K_10kpa',
       'Mix_Loading_C4H4S_363K_10kpa', 'Mix_Loading_C4H4S_363K_100kpa',
       'Mix_Loading_C6H6_298K_10kpa', 'Mix_Loading_C6H6_363K_10kpa',
       'Mix_Loading_C6H6_363K_100kpa', 'Selectivity_298K_10kpa',
       'Selectivity_363K_10kpa', 'Selectivity_298K_100kpa',
       'Working_capacity_PSA_363K_100kpa_10kpa',
       'Working_capacity_TSA_10kpa_298K_363K'],axis=1)
correlations=data.corr()
correction=abs(correlations)
fig,ax=plt.subplots(figsize=(20,20))
sns.heatmap(correlations,vmax=1.0,center=0,fmt='.2f',square=True,linewidths=.5,
            annot=True,cbar_kws={"shrink":.70})
# plt.show()

p1 = []
p2 = []
p3 = []
#Prepare Data
for k in range(len(Target)):
    # X=data.drop(['%s'%Target[k]],axis=1)
    X=data
    Feature_Descriptor.remove(Target[k])
    Y=data_1['%s'%Target[k]]
    scaler = StandardScaler(copy=True, with_mean=True, with_std=True).fit(X)
    X = scaler.transform(X)
    # Heat、Loading数据采用正则化
    normalizer = preprocessing.Normalizer(copy=True, norm='l2').fit(Y.values.reshape(1, -1))
    Y = normalizer.transform(Y.values.reshape(1, -1))
    Y = np.transpose(Y)
    #Divide Dataset
    X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.3, random_state=None)
    # print(X_train,"\n++++++++++",X_test,"\n+++++++++++",Y_train,"\n++++++++++++",Y_test)
    MAE = []
    MSE = []
    RMSE = []
    R2 = []
    MAPE = []
    xgboost = XGBRegressor(max_depth=1, learning_rate=0.1, n_estimators=100,silent=True, objective='reg:linear', booster='gbtree', n_jobs=1,
                           gamma=0, min_child_weight=1, max_delta_step=0, subsample=1, colsample_bytree=1,colsample_bylevel=1, reg_alpha=0,
                           reg_lambda=1, scale_pos_weight=1,base_score=0.5, random_state=0, missing=None)
    xgboost.fit(X_train, Y_train)
    Y_pred = xgboost.predict(X_test)
    print(Target[k], np.mean(cross_val_score(xgboost, X=X_test, y=Y_pred, cv=10)),
          np.std(cross_val_score(xgboost, X=X_test, y=Y_pred, cv=10)))
    p1.append(Target[k])
    p2.append(np.mean(cross_val_score(xgboost, X=X_test, y=Y_pred, cv=10)))
    p3.append(np.std(cross_val_score(xgboost, X=X_test, y=Y_pred, cv=10)))
    from openpyxl import load_workbook

    os.chdir('F:/MOFS_Extra_Data/ML_Graph/XGBoost_Graph')
    wb = load_workbook('1.xlsx')
    wb1 = wb.active
    for i in range(len(p1)):
        wb1.cell(i + 1, 1, p1[i])
        wb1.cell(i + 1, 2, p2[i])
        wb1.cell(i + 1, 3, p3[i])
    wb.save('1.xlsx')

    Y_train = list(Y_train)
    Y_test = list(Y_test)
    Y_pred = list(Y_pred)
    Y_test = np.array(Y_test)
    Y_pred = np.array(Y_pred)
    Y_test_mean = np.mean(Y_test)
    Y_pred_mean = np.mean(Y_pred)
    # 平均绝对误差(MAE)，MAE值越小，说明预测模型拥有更好的精确度
    # 均方误差(MSE)，误差的平方的期望值，MSE值越小，预测模型拥有更好的精确度
    # 平均绝对百分比误差(MAPE)，偏离，MAPE值越小，说明预测模型拥有更好的精确度
    def rmse(Y_test, Y_pred):  # RMSE
        error = []
        for i in range(len(Y_test)):
            error.append(Y_test[i] - Y_pred[i])
        squareError = []
        for i in error:
            squareError.append(i * i)
        return sqrt(sum(squareError) / len(squareError))
    c = mean_absolute_error(Y_test, Y_pred)
    d = mean_squared_error(Y_test, Y_pred)
    e = rmse(Y_test, Y_pred)
    f = r2_score(Y_test, Y_pred)
    # MAPE
    for i in range(len(Y_test)):
        if Y_test[i] == 0:
            Y_test[i] = Y_test_mean
    for i in range(len(Y_pred)):
        if Y_pred[i] == 0:
            Y_pred[i] = Y_pred_mean
    def mape(Y_test, Y_pred):
        n = len(Y_test)
        mape = sum(np.abs((Y_test - Y_pred) / Y_test)) / n * 100
        return mape
    g = mape(Y_test, Y_pred)
    MAE.append(c)
    MSE.append(d)
    RMSE.append(e)
    R2.append(f)
    MAPE.append(g)
    # 绘图
    fig = plt.figure()
    plt.scatter(Y_test, Y_pred)
    plt.xlabel("Y_test")
    plt.ylabel("Y_pred")
    plt.xlim(0, 0.1)
    plt.ylim(0, 0.1)
    os.chdir('F:\MOFS_Extra_Data\ML_Graph\XGBoost_Graph')
    plt.savefig('Y_test_predict_%s.png'%Target[k])
    # print(Target[k],"GBR:", "MAE:", np.mean(MAE), "MSE:", np.mean(MSE), "RMSE:", np.mean(RMSE),
    #       "R2_Score:", np.mean(R2), "MAPE:",np.mean(MAPE))

    importance=xgboost.feature_importances_
    print(len(importance))
    importance_1=[]
    for i in importance:
        if i != 0:
            importance_1.append(i)
    Impt_Series=pd.Series(importance_1)
    Impt_Series=Impt_Series.sort_values(ascending=True)
    Y=list(Impt_Series.index)
    print(Y)
    Y_index=[]
    for i in Y:
        Y_index.append(Feature_Descriptor[i])
    print(Y_index)
    plt.figure(figsize=(10,10))
    plt.barh(range(len(Y)),Impt_Series.values,tick_label=Y_index,color='steelblue')
    plt.xticks(fontsize=10)
    plt.yticks(fontsize=10)
    for y, x in enumerate(Impt_Series.values):
        plt.text(x + 0.0001, y, '%s' % round(x, 3), va='center')
    os.chdir('F:\MOFS_Extra_Data\ML_Graph\XGBoost_Graph')
    plt.savefig('Feature_Importance_%s.png' % Target[k])
    Feature_Descriptor.append(Target[k])
    '''
    # 特征重要性
    color = sns.color_palette()
    sns.set_style('darkgrid')
    feature_list = X_train
    feature_importance = GBR.feature_importances_
    sorted_idx = np.argsort(feature_importance)
    plt.figure(figsize=(5, 7))
    plt.barh(range(len(sorted_idx)), feature_importance[sorted_idx],
             align='center')
    plt.yticks(range(len(sorted_idx)), feature_list[sorted_idx])
    plt.xlabel('Importance')
    plt.title('Feature Importance')
    # plt.draw()
    # plt.show()
    os.chdir('F:\MOFS_Extra_Data\Test_1_Graph')
    plt.savefig('Feature_Importance_%s.png'%Target[k])
    '''

    with open('result.txt', 'a') as f:
        f.write(Target[k])
        f.write('\n')
        f.write("MAE: ")
        f.write(str(np.mean(MAE)))
        f.write('\t')
        f.write("MSE: ")
        f.write(str(np.mean(MSE)))
        f.write('\t')
        f.write("RMSE: ")
        f.write(str(np.mean(RMSE)))
        f.write('\t')
        f.write("R2_Score: ")
        f.write(str(np.mean(R2)))
        f.write('\t')
        f.write("MAPE: ")
        f.write(str(np.mean(MAPE)))
        f.write('\n')