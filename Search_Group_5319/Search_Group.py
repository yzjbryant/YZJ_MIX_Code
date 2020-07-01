import os
import re
import numpy as np
import math
from openpyxl import load_workbook
import xlrd

path_1='F:\MOFS_Extra_Data\cif\MOFs'
#/home/yinzj/Machine_Learning/MOFs

#获取文件名
file_name=os.listdir(path_1)
# print(file_name)
#读每个文件，获取元素
#1.获取行数
head_element=[]
tail_element=[]
for i in range(len(file_name)):
    with open(os.path.join(path_1,file_name[i]),'r') as f:
        count=0
        for line in f.readlines():
            count+=1
            if '_atom_site_occupancy' in line:
                head_element.append(count)
for i in range(len(file_name)):
    with open(os.path.join(path_1,file_name[i]),'r') as f:
        count=0  #获取
        for line in f.readlines():
            count+=1
            if '_geom_bond_atom_site_label_1' in line:
                tail_element.append(count-2)
# group_1_F=['0' for i in range(len(file_name))]
# group_2_Cl=['0' for i in range(len(file_name))]
# group_3_Br=['0' for i in range(len(file_name))]
# group_4_I=['0' for i in range(len(file_name))]
group_5_CH3=['0' for i in range(len(file_name))]
group_6_CH2_CH3=['0' for i in range(len(file_name))]
group_7_CH2_CH2_CH3=['0' for i in range(len(file_name))]
group_8_O=['0' for i in range(len(file_name))]
group_9_O_OH=['0' for i in range(len(file_name))]
group_10_OH=['0' for i in range(len(file_name))]
group_11_O_CH3=['0' for i in range(len(file_name))]
group_12_O_CH2_CH3=['0' for i in range(len(file_name))]
group_13_O_CH2_CH2_CH3=['0' for i in range(len(file_name))]
group_14_NH2=['0' for i in range(len(file_name))]
group_15_C_N=['0' for i in range(len(file_name))]
group_16_NH_CH3=['0' for i in range(len(file_name))]
group_17_N_O_O=['0' for i in range(len(file_name))]
group_18_C6H6=['0' for i in range(len(file_name))]
# group_19_S=['0' for i in range(len(file_name))]
#2.根据获取元素




'''
#先判断F,Cl,Br,I,S
for i in range(len(file_name)):
    init_element = []
    last_element = []
    with open(os.path.join(path_1,file_name[i]),'r') as f:
        for line in f.readlines()[head_element[i]:tail_element[i]]:
            number=[0,1,2,3,4,5,6,7,8,9]
            x=line[0:7]
            for u in number:
                x=re.sub('%s'%number[u],'',x)
                x=re.sub(' ','',x)
            init_element.append(x)
        [last_element.append(h) for h in init_element if not h in last_element]
        # print(last_element)
    #先判断F,Cl,Br,I,S
        element_1=['F','Cl','Br','I','S']
        for j in range(len(last_element)):
            if last_element[j]=='F':
                # print('1')
                group_1_F[i]='exist'
print(group_1_F)

for i in range(len(file_name)):
    init_element = []
    last_element = []
    with open(os.path.join(path_1, file_name[i]), 'r') as f:
        for line in f.readlines()[head_element[i]:tail_element[i]]:
            number = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
            x = line[0:7]
            for u in number:
                x = re.sub('%s' % number[u], '', x)
                x = re.sub(' ', '', x)
            init_element.append(x)
        [last_element.append(h) for h in init_element if not h in last_element]
        # 先判断F,Cl,Br,I,S
        element_1 = ['F', 'Cl', 'Br', 'I', 'S']
        for j in range(len(last_element)):
            if last_element[j] == 'Cl':
                group_2_Cl[i] = 'exist'

for i in range(len(file_name)):
    init_element = []
    last_element = []
    with open(os.path.join(path_1, file_name[i]), 'r') as f:
        for line in f.readlines()[head_element[i]:tail_element[i]]:
            number = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
            x = line[0:7]
            for u in number:
                x = re.sub('%s' % number[u], '', x)
                x = re.sub(' ', '', x)
            init_element.append(x)
        [last_element.append(h) for h in init_element if not h in last_element]
        # 先判断F,Cl,Br,I,S
        element_1 = ['F', 'Cl', 'Br', 'I', 'S']
        for j in range(len(last_element)):
            if last_element[j] == 'Br':
                group_3_Br[i] = 'exist'

for i in range(len(file_name)):
    init_element = []
    last_element = []
    with open(os.path.join(path_1, file_name[i]), 'r') as f:
        for line in f.readlines()[head_element[i]:tail_element[i]]:
            number = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
            x = line[0:7]
            for u in number:
                x = re.sub('%s' % number[u], '', x)
                x = re.sub(' ', '', x)
            init_element.append(x)
        [last_element.append(h) for h in init_element if not h in last_element]
        # 先判断F,Cl,Br,I,S
        element_1 = ['F', 'Cl', 'Br', 'I', 'S']
        for j in range(len(last_element)):
            if last_element[j] == 'I':
                group_4_I[i] = 'exist'

for i in range(len(file_name)):
    init_element = []
    last_element = []
    with open(os.path.join(path_1, file_name[i]), 'r') as f:
        for line in f.readlines()[head_element[i]:tail_element[i]]:
            number = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
            x = line[0:7]
            for u in number:
                x = re.sub('%s' % number[u], '', x)
                x = re.sub(' ', '', x)
            init_element.append(x)
        [last_element.append(h) for h in init_element if not h in last_element]
        # 先判断F,Cl,Br,I,S
        element_1 = ['F', 'Cl', 'Br', 'I', 'S']
        for j in range(len(last_element)):
            if last_element[j] == 'S':
                group_19_S[i] = 'exist'
'''

#读每个文件，获取键
#1.获取行数
head_bond=[]
tail_bond=[]
for i in range(len(file_name)):
    with open(os.path.join(path_1,file_name[i]),'r') as f:
        count=0
        for line in f.readlines():
            count+=1
            if '_ccdc_geom_bond_type' in line:
                head_bond.append(count)
for i in range(len(file_name)):
    with open(os.path.join(path_1,file_name[i]),'r') as f:
        count=0
        for line in f.readlines():
            count+=1
        tail_bond.append(count)

'''
#判断group-5 -CH3
for q in range(len(file_name)):
    first_column = []
    second_column = []
    bond=[]
    with open(os.path.join(path_1,file_name[q]),'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x=line[0:7]
            x=re.sub(' ','',x)
            y=line[7:14]
            y=re.sub(' ','',y)
            z=line[29:30]
            first_column.append(x)
            second_column.append(y)
            bond.append(z)
            m=[]
        for i in range(9,len(first_column)):
            if first_column[i-3]==first_column[i-2]==first_column[i-1]==first_column[i] and 'C' in first_column[i-3] and 'Cu' not in first_column[i-3] and 'Cl' not in first_column[i-3]:
                if ('H' in second_column[i-3] and 'H' in second_column[i-2]  and 'H' in second_column[i-1] and 'C' in second_column[i]) or \
                    ('H' in second_column[i - 3] and 'H' in second_column[i - 2] and 'C' in second_column[i - 1] and 'H' in second_column[i]) or \
                    ('H' in second_column[i - 3] and 'C' in second_column[i - 2] and 'H' in second_column[i - 1] and 'H' in second_column[i]) or \
                    ('C' in second_column[i - 3] and 'H' in second_column[i - 2] and 'H' in second_column[i - 1] and 'H' in second_column[i]):
                    m.append(second_column[i-3])
                    m.append(second_column[i - 2])
                    m.append(second_column[i - 1])
                    m.append(second_column[i])
                    for j in m:
                        if 'C' in j:
                            r=second_column.index(j)
                            if 'C' in first_column[r] and 'A' in bond[r]:
                                print('group-5')
                                group_5_CH3[q]='exist'
                    m = []
                    continue
        print(file_name[q])
        continue
'''
'''
#判断group-14 -NH2
for q in range(len(file_name)):
    first_column = []
    second_column = []
    with open(os.path.join(path_1, file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            first_column.append(x)
            second_column.append(y)
        for i in range(9, len(first_column)):
            if second_column[i - 1] == second_column[i] and 'N' in second_column[i-1]:
                    if 'H' in first_column[i - 1] and 'H' in first_column[i]:
                        x = re.sub('H', '', first_column[i - 1])
                        y = re.sub('H', '', first_column[i])
                        if abs(int(x) - int(y)) == 1:
                            print("have group-14")
                            group_14_NH2[q]='exist'
                            continue
        print(q)
        continue
'''
'''
#判断group-15 -C-N
for q in range(len(file_name)):
    first_column = []
    second_column = []
    bond=[]
    with open(os.path.join(path_1, file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            z=line[29:30]
            first_column.append(x)
            second_column.append(y)
            bond.append(z)
        for i in range(9,len(first_column)):
            if first_column[i-1]==first_column[i] and 'C' in first_column[i]:
                if ('C' in second_column[i-1] and 'N' in second_column[i] and 'T' in bond[i]) or ('N' in second_column[i-1] and 'C' in second_column[i] and 'T' in bond[i-1]):
                    print("have group-15")
                    group_15_C_N[q]='exist'
                    continue
        print(q)
        continue
'''
'''
#判断group-16 -NH-CH3
for q in range(len(file_name)):
    first_column = []
    second_column = []
    with open(os.path.join(path_1, file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            first_column.append(x)
            second_column.append(y)
        for i in range(9, len(first_column)):
            if (first_column[i - 3]==first_column[i - 2]==first_column[i - 1]==first_column[i] and 'C' in first_column[i-3]) and\
                            'Cu' not in first_column[i-3] and 'Cl' not in first_column[i-3]:
                if ('N' in second_column[i - 3] and 'H' in second_column[i - 2] and 'H' in second_column[i - 1] and 'H' in second_column[i]) or \
                        ('H' in second_column[i - 3] and 'N' in second_column[i - 2] and 'H' in second_column[i - 1] and 'H' in second_column[i]) or \
                        ('H' in second_column[i - 3] and 'H' in second_column[i - 2] and 'N' in second_column[i - 1] and 'H' in second_column[i]) or \
                        ('H' in second_column[i - 3] and 'H' in second_column[i - 2] and 'H' in second_column[i - 1] and 'N' in second_column[i]):
                    print('have group-16')
                    group_16_NH_CH3[q]='exist'
                    continue
        print(q)
        continue
'''
'''
#判断group-17 -N-O-O
for q in range(len(file_name)):
    first_column = []
    second_column = []
    with open(os.path.join(path_1, file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            first_column.append(x)
            second_column.append(y)
        for i in range(9, len(first_column)):
            if first_column[i - 1] == first_column[i] and 'N' in first_column[i-1]:
                if 'O' in second_column[i - 1] and 'O' in second_column[i]:
                    x = re.sub('O', '', second_column[i - 1])
                    y = re.sub('O', '', second_column[i])
                    if abs(int(x) - int(y)) == 1:
                        print('have group-17')
                        group_17_N_O_O[q]='exist'
        print(q)
        continue
'''
'''
#判断group-18 -C6H6
for q in range(len(file_name)):
    first_column = []
    second_column = []
    with open(os.path.join(path_1, file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            first_column.append(x)
            second_column.append(y)
        for i in range(11,len(first_column)):
            if first_column[i-11]==first_column[i-10]==first_column[i-9] and \
                first_column[i-8]==first_column[i-7] and \
                first_column[i-6]==first_column[i-5] and \
                first_column[i-4]==first_column[i-3] and \
                first_column[i-2]==first_column[i-1] and \
                'C' in first_column[i-11] and 'C' in first_column[i-8] and \
                'C' in first_column[i-6] and 'C' in first_column[i-4] and \
                'C' in first_column[i-2] and 'C' in first_column[i] and \
                'Cu' not in first_column[i - 11] and 'Cu' not in first_column[i - 8] and \
                'Cu' not in first_column[i - 6] and 'Cu' not in first_column[i - 4] and \
                'Cu' not in first_column[i - 2] and 'Cu' not in first_column[i] and \
                'Cl' not in first_column[i - 11] and 'Cl' not in first_column[i - 8] and \
                'Cl' not in first_column[i - 6] and 'Cl' not in first_column[i - 4] and \
                'Cl' not in first_column[i - 2] and 'Cl' not in first_column[i]:
                x=re.sub('C','',first_column[i-11])
                y=re.sub('C','',first_column[i-8])
                z=re.sub('C','',first_column[i-6])
                h=re.sub('C','',first_column[i-4])
                u=re.sub('C','',first_column[i-2])
                m=re.sub('C','',first_column[i])
                if abs(int(x)-int(y))==1 and abs(int(y)-int(z))==1 and abs(int(z)-int(h)) and \
                    abs(int(h)-int(u))==1 and abs(int(u)-int(m))==1 and abs(int(x)-int(m))==5:
                    if 'C' in second_column[i-11] and 'C' in second_column[i-10] and 'C' in second_column[i-9]:
                        if (('C' in second_column[i-8] and 'H' in second_column[i-7]) or ('H' in second_column[i-8] and 'C' in second_column[i-7])) and \
                            (('C' in second_column[i-6] and 'H' in second_column[i-5]) or ('H' in second_column[i-6] and 'C' in second_column[i-5])) and \
                            (('C' in second_column[i-4] and 'H' in second_column[i-3]) or ('H' in second_column[i-4] and 'C' in second_column[i-3])) and \
                            (('C' in second_column[i-2] and 'H' in second_column[i-1]) or ('H' in second_column[i-2] and 'C' in second_column[i-1])) and \
                            ('H' in second_column[i]):
                            print('have group-18')
                            group_18_C6H6[q] = 'exist'
                            continue
        print(q)
        continue
'''
'''
#判断group-6 -CH2-CH3
for q in range(len(file_name)):
    first_column = []
    second_column = []
    bond=[]
    with open(os.path.join(path_1, file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            z=line[29:30]
            first_column.append(x)
            second_column.append(y)
            bond.append(z)
        for i in range(5, len(first_column)):
            if (first_column[i-5]==first_column[i-4]==first_column[i-3] and 'C' in first_column[i-5] and 'Cu' not in first_column[i-5] and 'Cl' not in first_column[i-5]) and \
                    (first_column[i-2]==first_column[i-1]==first_column[i] and 'C' in first_column[i-2] and 'Cu' not in first_column[i-2] and 'Cl' not in first_column[i-2]):
                x=re.sub('C','',first_column[i-5])
                y=re.sub('C','',first_column[i-2])
                if abs(int(x)-int(y))==1:
                    if ('C' in second_column[i-5] and 'H' in second_column[i-4] and 'H' in second_column[i-3]) or \
                            ('H' in second_column[i - 5] and 'C' in second_column[i - 4] and 'H' in second_column[i - 3]) or \
                            ('H' in second_column[i - 5] and 'H' in second_column[i - 4] and 'C' in second_column[i - 3]) and \
                            ('H' in second_column[i-2] and 'H' in second_column[i-1] and 'H' in second_column[i]):
                        if first_column[i-2]==second_column[i-5] or first_column[i-2]==second_column[i-4] or first_column[i-2]==second_column[i-3]:
                            if first_column[i-5] in second_column:
                                r=second_column.index(first_column[i-5])
                                if r<len(first_column)-5:
                                    if (first_column[r]==first_column[r+1]==first_column[r+2] and 'C' in first_column[r] and 'Cu' not in first_column[r] and 'Cl' not in first_column[r] and ('A' in bond[r] or 'A' in bond[r+1] or 'A' in bond[r+2])) or \
                                            (first_column[r-1]==first_column[r]==first_column[r] and 'C' in first_column[r+1] and 'Cu' not in first_column[r+1] and 'Cl' not in first_column[r+1] and ('A' in bond[r-1] or 'A' in bond[r] or 'A' in bond[r+1])) or \
                                            (first_column[r - 2] == first_column[r] == first_column[r-1] and 'C' in first_column[r] and 'Cu' not in first_column[r] and 'Cl' not in first_column[r] and ('A' in first_column[r-2] or 'A' in first_column[r-1] or 'A' in first_column[r])):
                                            print('group-6')
                                            group_6_CH2_CH3[q]='exist'
                                            continue

        print(file_name[q])
        continue
'''
'''
#判断group-7 -CH2-CH2-CH3
for q in range(len(file_name)):
    first_column = []
    second_column = []
    with open(os.path.join(path_1, file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            first_column.append(x)
            second_column.append(y)
        for i in range(9, len(first_column)):
            if (first_column[i - 9] == first_column[i - 8] == first_column[i - 7] == first_column[i - 6] and 'C' in first_column[i-9] and 'Cu' not in first_column[i-9] and 'Cl' not in first_column[i-9]) and \
                    (first_column[i - 5] == first_column[i - 4] == first_column[i - 3] and 'C' in first_column[i-5] and 'Cu' not in first_column[i-5] and 'Cl' not in first_column[i-5]) and \
                    (first_column[i - 2] == first_column[i - 1] == first_column[i] and 'C' in first_column[i-2] and 'Cu' not in first_column[i-2] and 'Cl' not in first_column[i-2]):
                if ('C' in second_column[i-9] and 'C' in second_column[i-8] and 'H' in second_column[i-7] and 'H' in second_column[i-6]) or \
                    ('C' in second_column[i - 9] and 'H' in second_column[i - 8] and 'H' in second_column[i - 7] and 'C' in second_column[i - 6]) or \
                    ('C' in second_column[i - 9] and 'H' in second_column[i - 8] and 'C' in second_column[i - 7] and 'H' in second_column[i - 6]) or \
                    ('H' in second_column[i - 9] and 'H' in second_column[i - 8] and 'C' in second_column[i - 7] and 'C' in second_column[i - 6]) or \
                    ('H' in second_column[i - 9] and 'C' in second_column[i - 8] and 'H' in second_column[i - 7] and 'C' in second_column[i - 6]) or \
                    ('H' in second_column[i - 9] and 'C' in second_column[i - 8] and 'C' in second_column[i - 7] and 'H' in second_column[i - 6]):
                    if ('H' in second_column[i-5] and 'H' in second_column[i-4] and 'C' in second_column[i-3]) or \
                        ('H' in second_column[i - 5] and 'C' in second_column[i - 4] and 'H' in second_column[i - 3]) or \
                        ('C' in second_column[i - 5] and 'H' in second_column[i - 4] and 'H' in second_column[i - 3]):
                        if 'H' in second_column[i-2] and 'H' in second_column[i-1] and 'H' in second_column[i]:
                            print('have group-7')
                            group_7_CH2_CH2_CH3[q] = 'exist'
                            continue
        print(q)
        continue
'''
''' 
#判断group-8 -O
for q in range(len(file_name)):
    first_column = []
    second_column = []
    bond=[]
    with open(os.path.join(path_1, file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            z=line[29:30]
            first_column.append(x)
            second_column.append(y)
            bond.append(z)
        for i in range(9, len(first_column)):
            if first_column[i - 2] == first_column[i-1]==first_column[i]:
                if 'C' in first_column[i - 2] and 'C' in first_column[i - 1] and 'C' in first_column[i] and \
                                'Cl' not in first_column[i-2] and 'Cl' not in first_column[i-1] and 'Cl' not in first_column[i] and \
                                'Cu' not in first_column[i-2] and'Cu' not in first_column[i-1] and 'Cu' not in first_column[i]:
                    if 'O' in second_column[i - 2] and 'C' in second_column[i-1] and 'H' in second_column[i] or \
                        'O' in second_column[i - 2] and 'H' in second_column[i - 1] and 'C' in second_column[i] or \
                        'C' in second_column[i - 2] and 'H' in second_column[i - 1] and 'O' in second_column[i] or \
                        'C' in second_column[i - 2] and 'O' in second_column[i - 1] and 'H' in second_column[i] or \
                        'H' in second_column[i - 2] and 'O' in second_column[i - 1] and 'C' in second_column[i] or \
                        'H' in second_column[i - 2] and 'C' in second_column[i - 1] and 'O' in second_column[i]:
                        if 'D' in bond[i - 2] or 'D' in bond[i - 1] or 'D' in bond[i]:  # D 表示双键
                            print('have group-8')
                            group_8_O[q]='exist'
                            continue
        print(q)
        continue
'''
'''
#判断group-9 -O-OH
for q in range(len(file_name)):
    first_column = []
    second_column = []
    bond=[]
    with open(os.path.join(path_1, file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            z=line[29:30]
            first_column.append(x)
            second_column.append(y)
            bond.append(z)
        for i in range(9, len(first_column)):
            if first_column[i - 1] == first_column[i] and 'C' in first_column[i]:   #D为双键
                if 'O' in second_column[i - 1] and 'O' in second_column[i]:
                    x=re.sub('O','',second_column[i-1])
                    y=re.sub('O','',second_column[i])
                    if abs(int(x)-int(y))==1:
                        if 'D' in bond[i-1] or 'D' in bond[i]:
                            print('have group-9')
                            group_9_O_OH[q]='exist'
                            continue
        print(q)
        continue
'''
'''
#判断group-10 -OH
for q in range(len(file_name)):
    first_column = []
    second_column = []
    bond=[]
    with open(os.path.join(path_1,file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            z=line[29:30]
            first_column.append(x)
            second_column.append(y)
            bond.append(z)
        for i in range(9,len(first_column)):
            if 'H' in first_column[i] and 'O' in second_column[i]:
                r=second_column.index(second_column[i])
                if r < len(first_column)-5:
                    if first_column[r] == first_column[r+1] and 'C' in first_column[r] and 'Cu' not in first_column[r] and 'Cl' not in first_column[r]:
                        if 'C' in second_column[r+1] and 'A' in bond[r+1]:
                            print('have group-10')
                            group_10_OH[q] = 'exist'
                            continue
        print(file_name[q])
        continue
'''
'''
#判断group-11 -O-CH3
for q in range(len(file_name)):
    first_column = []
    second_column = []
    with open(os.path.join(path_1, file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            first_column.append(x)
            second_column.append(y)
        for i in range(9, len(first_column)):
            for j in range(3, len(first_column)):
                if first_column[i - 2] == first_column[i - 1] == first_column[i] and 'C' in first_column[i - 2] and 'C' in \
                        first_column[i - 1] and 'C' in first_column[i] and 'Cu' not in first_column[i - 2] and 'Cu' not in first_column[i - 1] and 'Cu' not in first_column[i] and 'Cl' not in first_column[i - 2] and 'Cl' not in first_column[i - 1] and 'Cl' not in first_column[i]:
                    if 'C' in second_column[i - 2] and 'O' in second_column[i - 1] and 'C' in second_column[i]:
                        if first_column[j - 3] == first_column[j - 2] == first_column[j - 1] == first_column[j]:
                            if 'C' in first_column[j - 3] and 'C' in first_column[j - 2] and 'C' in first_column[j - 1] and 'C' in first_column[j] and 'Cl' not in \
                    first_column[j-3] and 'Cl' not in first_column[j-2] and 'Cl' not in first_column[j-1] and 'Cl' not in first_column[j] and 'Cu' not in \
                    first_column[j-3] and 'Cu' not in first_column[j-2] and 'Cu' not in first_column[j-1] and 'Cu' not in first_column[j]:
                                if 'H' in second_column[j - 3] and 'O' in second_column[j - 2] and 'H' in second_column[j - 1] and 'H' in second_column[j]:
                                    print('have group-11')
                                    group_11_O_CH3[q]='exist'
                                    continue
            continue
        print(q)
        continue
'''
'''
#判断group-12 -O-CH2-CH3
for q in range(len(file_name)):
    first_column = []
    second_column = []
    bond=[]
    with open(os.path.join(path_1, file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            z=line[29:30]
            first_column.append(x)
            second_column.append(y)
            bond.append(z)
            m=[]
        for i in range(9,len(first_column)):
            if (first_column[i-6]==first_column[i-5]==first_column[i-4]==first_column[i-3] and 'C' in first_column[i-3] and 'Cl' not in first_column[i-3] and 'Cu' not in first_column[i-3]) and \
                    (first_column[i-2]==first_column[i-1]==first_column[i] and 'C' in first_column[i-1] and 'Cl' not in first_column[i-1] and 'Cu' not in first_column[i-1]):
                if ('C' in second_column[i-6] and 'O' in second_column[i-5] and 'H' in second_column[i-4] and 'H' in second_column[i-3]) or \
                    ('C' in second_column[i - 6] and 'H' in second_column[i - 5] and 'O' in second_column[i - 4] and 'H' in second_column[i - 3]) or \
                    ('C' in second_column[i - 6] and 'H' in second_column[i - 5] and 'H' in second_column[i - 4] and 'O' in second_column[i - 3]) or \
                    ('H' in second_column[i - 6] and 'H' in second_column[i - 5] and 'O' in second_column[i - 4] and 'C' in second_column[i - 3]) or \
                    ('H' in second_column[i - 6] and 'H' in second_column[i - 5] and 'C' in second_column[i - 4] and 'O' in second_column[i - 3]) or \
                    ('H' in second_column[i - 6] and 'C' in second_column[i - 5] and 'O' in second_column[i - 4] and 'H' in second_column[i - 3]) or \
                    ('H' in second_column[i - 6] and 'C' in second_column[i - 5] and 'H' in second_column[i - 4] and 'O' in second_column[i - 3]) or \
                    ('H' in second_column[i - 6] and 'O' in second_column[i - 5] and 'C' in second_column[i - 4] and 'H' in second_column[i - 3]) or \
                    ('H' in second_column[i - 6] and 'O' in second_column[i - 5] and 'H' in second_column[i - 4] and 'C' in second_column[i - 3]) or \
                    ('O' in second_column[i - 6] and 'H' in second_column[i - 5] and 'H' in second_column[i - 4] and 'C' in second_column[i - 3]) or \
                    ('O' in second_column[i - 6] and 'H' in second_column[i - 5] and 'C' in second_column[i - 4] and 'H' in second_column[i - 3]) or \
                    ('O' in second_column[i - 6] and 'C' in second_column[i - 5] and 'H' in second_column[i - 4] and 'H' in second_column[i - 3]):
                    if 'H' in second_column[i-2] and 'H' in second_column[i-1] and 'H' in second_column[i]:
                        print("have group-12")
                        group_12_O_CH2_CH3[q] = 'exist'
                        continue
        print(q)
        continue
'''
'''
#判断group-13 -O-CH2-CH2-CH3
for q in range(len(file_name)):
    first_column = []
    second_column = []
    with open(os.path.join(path_1, file_name[q]), 'r') as f:
        for line in f.readlines()[head_bond[q]:tail_bond[q]]:
            x = line[0:7]
            x = re.sub(' ', '', x)
            y = line[7:14]
            y = re.sub(' ', '', y)
            first_column.append(x)
            second_column.append(y)
        for i in range(9,len(first_column)):
            if first_column[i-9]==first_column[i-8]==first_column[i-7]==first_column[i-6] and \
               first_column[i - 5] == first_column[i - 4] == first_column[i - 3] and \
               first_column[i - 2] == first_column[i - 1] == first_column[i]:
                if 'C' in first_column[i-9] and 'Cl' not in first_column[i-9] and 'Cu' not in first_column[i-9] and \
                   'C' in first_column[i - 5] and 'Cl' not in first_column[i - 5] and 'Cu' not in first_column[i - 5] and \
                   'C' in first_column[i - 2] and 'Cl' not in first_column[i - 2] and 'Cu' not in first_column[i - 2]:
                    x = re.sub('C', '', first_column[i - 9])
                    y = re.sub('C', '', first_column[i - 5])
                    z = re.sub('C', '', first_column[i - 2])
                    if abs(int(x)-int(y))==1 and abs(int(x)-int(z))==2 and abs(int(y)-int(z))==1:
                        if 'O' in second_column[i-9] and 'H' in second_column[i-8] and 'C' in second_column[i-7] and 'H' in second_column[i-6] or \
                            'O' in second_column[i-9] and 'C' in second_column[i-8] and 'H' in second_column[i-7] and 'H' in second_column[i-6] or \
                            'O' in second_column[i - 9] and 'H' in second_column[i - 8] and 'H' in second_column[i - 7] and 'C' in second_column[i - 6] or \
                            'C' in second_column[i - 9] and 'H' in second_column[i - 8] and 'H' in second_column[i - 7] and 'O' in second_column[i - 6] or \
                            'C' in second_column[i - 9] and 'H' in second_column[i - 8] and 'O' in second_column[i - 7] and 'H' in second_column[i - 6] or \
                            'C' in second_column[i - 9] and 'O' in second_column[i - 8] and 'H' in second_column[i - 7] and 'H' in second_column[i - 6] or \
                            'H' in second_column[i - 9] and 'H' in second_column[i - 8] and 'C' in second_column[i - 7] and 'O' in second_column[i - 6] or \
                            'H' in second_column[i - 9] and 'H' in second_column[i - 8] and 'O' in second_column[i - 7] and 'C' in second_column[i - 6] or \
                            'H' in second_column[i - 9] and 'O' in second_column[i - 8] and 'H' in second_column[i - 7] and 'C' in second_column[i - 6] or \
                            'H' in second_column[i - 9] and 'O' in second_column[i - 8] and 'C' in second_column[i - 7] and 'H' in second_column[i - 6] or \
                            'H' in second_column[i - 9] and 'C' in second_column[i - 8] and 'H' in second_column[i - 7] and 'O' in second_column[i - 6] or \
                            'H' in second_column[i - 9] and 'C' in second_column[i - 8] and 'O' in second_column[i - 7] and 'H' in second_column[i - 6]:
                                if 'C' in second_column[i - 5] and 'H' in second_column[i - 4] and 'H' in second_column[i-3] or \
                                'H' in second_column[i - 5] and 'C' in second_column[i - 4] and 'H' in second_column[i-3] or \
                                'H' in second_column[i - 5] and 'H' in second_column[i - 4] and 'C' in second_column[i-3]:
                                    if 'H' in second_column[i - 2] and 'H' in second_column[i - 1] and 'H' in second_column[i]:
                                        print('have group-13')
                                        group_13_O_CH2_CH2_CH3[q] = 'exist'
                                        continue
        print(q)
        continue
'''

wb1=load_workbook('group.xlsx')
wb2=wb1.active
for i in range(len(file_name)):
    # wb2.cell(i+2,1,group_1_F[i])
    # wb2.cell(i + 2, 2, group_2_Cl[i])
    # wb2.cell(i + 2, 3, group_3_Br[i])
    # wb2.cell(i + 2, 4, group_4_I[i])
    # wb2.cell(i + 2, 5, group_5_CH3[i])
    # wb2.cell(i + 2, 6, group_6_CH2_CH3[i])
    # wb2.cell(i + 2, 7, group_7_CH2_CH2_CH3[i])
    # wb2.cell(i + 2, 8, group_8_O[i])
    # wb2.cell(i + 2, 9, group_9_O_OH[i])
    # wb2.cell(i + 2, 10, group_10_OH[i])
    # wb2.cell(i + 2, 11, group_11_O_CH3[i])
    # wb2.cell(i + 2, 12, group_12_O_CH2_CH3[i])
    # wb2.cell(i + 2, 13, group_13_O_CH2_CH2_CH3[i])
    # wb2.cell(i + 2, 14, group_14_NH2[i])
    # wb2.cell(i + 2, 15, group_15_C_N[i])
    # wb2.cell(i + 2, 16, group_16_NH_CH3[i])
    # wb2.cell(i + 2, 17, group_17_N_O_O[i])
    # wb2.cell(i + 2, 18, group_18_C6H6[i])
    # wb2.cell(i + 2, 19, group_19_S[i])
    wb2.cell(1,20,'Material')
    wb2.cell(i + 2, 20, file_name[i])
wb1.save('group.xlsx')





