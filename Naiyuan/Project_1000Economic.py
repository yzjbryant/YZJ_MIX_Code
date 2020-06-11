import os
import pandas as pd
import docx
import xlrd
from docx import Document
from xlutils.copy import copy
import openpyxl
from openpyxl import load_workbook
import re

#读数据
path1='F:\江苏省研究生教育教学改革研究与实践课题\奈园\五月——5.11——43个指标\全球前1000名世界经济学家\原始数据'
os.chdir(path1)
wb=xlrd.open_workbook('2016-2020.xlsx')
sheet1=wb.sheet_by_name('2016')
sheet2=wb.sheet_by_name('2017')
sheet3=wb.sheet_by_name('2018')
sheet4=wb.sheet_by_name('2019')
sheet5=wb.sheet_by_name('2020')

#2016
first_column_2016=[]
second_column_2016=[]
for i in range(sheet1.nrows):
    cells=sheet1.row_values(i)
    data1=str(cells[0])
    data2=str(cells[1])
    first_column_2016.append(data1)
    second_column_2016.append(data2)
first_column_except_0_2016=[]
for i in range(len(first_column_2016)):
    if first_column_2016[i] != '':
        first_column_except_0_2016.append(first_column_2016[i])
first_column_index_2016=[]
for i in range(len(first_column_except_0_2016)):
    first_column_index_2016.append(first_column_2016.index(first_column_except_0_2016[i]))
print(len(first_column_index_2016))
#姓名&学校
People_name_2016=[]
for i in range(len(first_column_index_2016)):
    People_name_2016.append(second_column_2016[first_column_index_2016[i]])
All_school_2016=[i for i in second_column_2016 if i not in People_name_2016]
# print(People_name_2016,len(People_name_2016))
print(len(All_school_2016))


























#2017
first_column_2017=[]
second_column_2017=[]
for i in range(sheet2.nrows):
    cells=sheet2.row_values(i)
    data1=str(cells[0])
    data2=str(cells[1])
    first_column_2017.append(data1)
    second_column_2017.append(data2)
first_column_except_0_2017=[]
for i in range(len(first_column_2017)):
    if first_column_2017[i] != '':
        first_column_except_0_2017.append(first_column_2017[i])
first_column_index_2017=[]
for i in range(len(first_column_except_0_2017)):
    first_column_index_2017.append(first_column_2017.index(first_column_except_0_2017[i]))
#姓名&学校
People_name_2017=[]
for i in range(len(first_column_index_2017)):
    People_name_2017.append(second_column_2017[first_column_index_2017[i]])
All_school_2017=[i for i in second_column_2017 if i not in People_name_2017]

#2018
first_column_2018=[]
second_column_2018=[]
for i in range(sheet3.nrows):
    cells=sheet3.row_values(i)
    data1=str(cells[0])
    data2=str(cells[1])
    first_column_2018.append(data1)
    second_column_2018.append(data2)
first_column_except_0_2018=[]
for i in range(len(first_column_2018)):
    if first_column_2018[i] != '':
        first_column_except_0_2018.append(first_column_2018[i])
first_column_index_2018=[]
for i in range(len(first_column_except_0_2018)):
    first_column_index_2018.append(first_column_2018.index(first_column_except_0_2018[i]))
#姓名&学校
People_name_2018=[]
for i in range(len(first_column_index_2018)):
    People_name_2018.append(second_column_2018[first_column_index_2018[i]])
All_school_2018=[i for i in second_column_2018 if i not in People_name_2018]

#2019
first_column_2019=[]
second_column_2019=[]
for i in range(sheet4.nrows):
    cells=sheet4.row_values(i)
    data1=str(cells[0])
    data2=str(cells[1])
    first_column_2019.append(data1)
    second_column_2019.append(data2)
first_column_except_0_2019=[]
for i in range(len(first_column_2019)):
    if first_column_2019[i] != '':
        first_column_except_0_2019.append(first_column_2019[i])
first_column_index_2019=[]
for i in range(len(first_column_except_0_2019)):
    first_column_index_2019.append(first_column_2019.index(first_column_except_0_2019[i]))
#姓名&学校
People_name_2019=[]
for i in range(len(first_column_index_2019)):
    People_name_2019.append(second_column_2019[first_column_index_2019[i]])
All_school_2019=[i for i in second_column_2019 if i not in People_name_2019]

#2020
first_column_2020=[]
second_column_2020=[]
for i in range(sheet5.nrows):
    cells=sheet5.row_values(i)
    data1=str(cells[0])
    data2=str(cells[1])
    first_column_2020.append(data1)
    second_column_2020.append(data2)
first_column_except_0_2020=[]
for i in range(len(first_column_2020)):
    if first_column_2020[i] != '':
        first_column_except_0_2020.append(first_column_2020[i])
first_column_index_2020=[]
for i in range(len(first_column_except_0_2020)):
    first_column_index_2020.append(first_column_2020.index(first_column_except_0_2020[i]))
#姓名&学校
People_name_2020=[]
for i in range(len(first_column_index_2020)):
    People_name_2020.append(second_column_2020[first_column_index_2020[i]])
All_school_2020=[i for i in second_column_2020 if i not in People_name_2020]
# print(len(All_school_2016))
# print(len(All_school_2017))
# print(len(All_school_2018))
# print(len(All_school_2019))
# print(len(All_school_2020))

# count=0
# for i in range(len(All_school_2016)):
#     if len(All_school_2016[i])<26:
#         count+=1
#         print(All_school_2016[i])
# print(count)



# location_0=[]
# location_1=[]
# location_2=[]
# for i in range(1,len(first_column_index)):
#     if first_column_index[i]-first_column_index[i-1]==1:
#         location_0.append(first_column_index[i-1])
#     if first_column_index[i]-first_column_index[i-1]==2:
#         location_1.append(first_column_index[i-1])
#     if first_column_index[i]-first_column_index[i-1]==3:
#         location_2.append(first_column_index[i-1])

# for i in range(2,len(second_column)+2):
#     for j in range(len(People_name_2016)):
#         if second_column[i-2]==People_name_2016[j]:
#             if second_column[i-1]!=People_name_2016[j]:
#                 First_school.append('1')
#             else:
#                 First_school.append('2')


#写文件
path2='F:\江苏省研究生教育教学改革研究与实践课题\奈园\五月——5.11——43个指标\全球前1000名世界经济学家\处理结果'
os.chdir(path2)
wb1=load_workbook('全球前1000名世界经济学家.xlsx')
wb2=wb1.active
for i in range(len(People_name_2016)):
    wb2.cell(i+2,1,People_name_2016[i])
    wb2.cell(i + 1002, 1, People_name_2017[i])
    wb2.cell(i + 2002, 1, People_name_2018[i])
    wb2.cell(i + 3002, 1, People_name_2019[i])
    wb2.cell(i + 4002, 1, People_name_2020[i])
wb1.save('全球前1000名世界经济学家.xlsx')