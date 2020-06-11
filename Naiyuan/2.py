import requests
import re
import lxml
import os
from lxml import etree
from openpyxl import load_workbook
import xlrd

headers={'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36'}
response=requests.get('http://www.gx211.com/gxmd/gx-bj.html',headers=headers)
html=response.text
url=re.findall('href="(.*?)"',html)
Name_of_Province=['北京','天津','河北','山西','辽宁','吉林','黑龙江','上海','江苏','浙江','安徽','福建','江西',
                '山东','河南','湖北','湖南','广东','内蒙古','广西','海南','重庆','四川','贵州','云南','西藏',
                '陕西','甘肃','青海','宁夏','新疆','港澳台']
Province_34_website=[]
for i in url:
    if 'gx-' in i:
        Province_34_website.append('http://www.gx211.com/gxmd/%s'%i)
Province_34_website.append('http://www.gx211.com/gxmd/gx-gat.html')
# print(Province_34_website)
All_website=[]
#爬取全国34个省网址
for i in range(len(Province_34_website)-1):
    response=requests.get(Province_34_website[i],headers=headers)
    url=response.text
    code=re.findall('charset=(.*?)"',url)
    html=response.content.decode(code[0])
    e_html=etree.HTML(html)
    uniersity_name=e_html.xpath('//a/text()')
    # print(uniersity_name)
    All_website.append(uniersity_name)
#港澳台——成功
response_gat=requests.get(Province_34_website[-1],headers=headers)
response_gat.encoding="gb2312"
html=response_gat.text
url=re.findall('<a href="(.*?)" target="_blank">',html)
# print("港澳台：",url)
All_website.append(url)
All_website_new=[]
for i in All_website:
    All_website_new+=i
print(len(All_website_new),All_website_new)

wb=load_workbook('科教文章库爬取链接地址.xlsx')
wb1=wb.active
for i in range(len(Name_of_Province)):
    wb1.cell(i+1,2,Name_of_Province[i])
# for i in range(len(All_website_new)):
#     wb1.cell(i+1,1,All_website_new[i])
wb.save('科教文章库爬取链接地址.xlsx')