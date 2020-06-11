import os
import pandas as pd
import docx
import xlrd
from docx import Document
from xlutils.copy import copy
import openpyxl
from openpyxl import load_workbook
import re

'''
#2014-2015
path_2014_2015_read_country='F:\奈园代码\本专科生国家奖学金\原始数据\p2014—2015学年度本专科生国家奖学金\中央高校'
path_2014_2015_read_city='F:\奈园代码\本专科生国家奖学金\原始数据\p2014—2015学年度本专科生国家奖学金\地方高校'
path_2014_2015_write='F:\奈园代码\本专科生国家奖学金\处理数据'
filename_2014_2015_country=os.listdir(path_2014_2015_read_country)
filename_2014_2015_city=os.listdir(path_2014_2015_read_city)

people_name_2014_2015=[]
school_2014_2015=[]
department_2014_2015=[]
project_2014_2015=[]
code_2014_2015=[]
gender_2014_2015=[]
nation_2014_2015=[]
date_2014_2015=[]
book_2014_2015=[]
#地区/中央
class_2014_2015=[]
class_2014_2015_country_index=[]
class_2014_2015_city_index=[]
class_2014_2015_country_index_name=[]
class_2014_2015_city_index_name=[]
for i in range(len(filename_2014_2015_country)):
    os.chdir(path_2014_2015_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2014_2015_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    for j in range(2,row_count):
        class_2014_2015_country_index.append(sheet.cell(j,1).value)
for i in range(len(class_2014_2015_country_index)):
    class_2014_2015_country_index_name.append('中央高校')
for i in range(len(filename_2014_2015_city)):
    os.chdir(path_2014_2015_read_city)
    workbook = xlrd.open_workbook(r'%s' % filename_2014_2015_city[i])
    sheet = workbook.sheet_by_index(0)
    row_count = sheet.nrows
    col_count = sheet.ncols
    for j in range(2, row_count):
        class_2014_2015_city_index.append(sheet.cell(j, 1).value)
for i in range(len(class_2014_2015_city_index)):
    class_2014_2015_city_index_name.append('地方高校')
class_2014_2015=class_2014_2015_country_index_name+class_2014_2015_city_index_name
#加地区
area_2014_2015=[]
area_2014_2015_index_country=[]
area_2014_2015_index_city=[]
area_2014_2015_index_country_name=[]
area_2014_2015_index_city_name=[]
for i in range(len(filename_2014_2015_country)):
    os.chdir(path_2014_2015_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2014_2015_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    area_2014_2015_index_country.append(row_count-2)
for i in range(len(area_2014_2015_index_country)):
    for j in range(area_2014_2015_index_country[i]):
        area_2014_2015_index_country_name.append('')
for i in range(len(filename_2014_2015_city)):
    os.chdir(path_2014_2015_read_city)
    workbook=xlrd.open_workbook(r'%s'%filename_2014_2015_city[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    area_2014_2015_index_city.append(row_count-2)
for i in range(len(area_2014_2015_index_city)):
    for j in range(area_2014_2015_index_city[i]):
        area_2014_2015_index_city_name.append(filename_2014_2015_city[i])
area_2014_2015=area_2014_2015_index_country_name+area_2014_2015_index_city_name


#读全部文件
for i in range(len(filename_2014_2015_country)):
    os.chdir(path_2014_2015_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2014_2015_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    for j in range(2,row_count):
        people_name_2014_2015.append(sheet.cell(j,1).value)
        school_2014_2015.append(sheet.cell(j, 2).value)
        department_2014_2015.append(sheet.cell(j, 3).value)
        project_2014_2015.append(sheet.cell(j, 4).value)
        code_2014_2015.append(sheet.cell(j, 5).value)
        gender_2014_2015.append(sheet.cell(j, 6).value)
        nation_2014_2015.append(sheet.cell(j, 7).value)
        date_2014_2015.append(sheet.cell(j, 8).value)
        book_2014_2015.append(sheet.cell(j, 9).value)
for i in range(len(filename_2014_2015_city)):
    os.chdir(path_2014_2015_read_city)
    workbook=xlrd.open_workbook(r'%s'%filename_2014_2015_city[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    for j in range(2,row_count):
        people_name_2014_2015.append(sheet.cell(j,1).value)
        school_2014_2015.append(sheet.cell(j, 2).value)
        department_2014_2015.append(sheet.cell(j, 3).value)
        project_2014_2015.append(sheet.cell(j, 4).value)
        code_2014_2015.append(sheet.cell(j, 5).value)
        gender_2014_2015.append(sheet.cell(j, 6).value)
        nation_2014_2015.append(sheet.cell(j, 7).value)
        date_2014_2015.append(sheet.cell(j, 8).value)
        book_2014_2015.append(sheet.cell(j, 9).value)

#写

os.chdir(path_2014_2015_write)
wb=load_workbook("2014—2015学年度本专科生国家奖学金.xlsx")
wb1=wb.active
for i in range(len(people_name_2014_2015)):
    wb1.cell(i+2,1,i+1)
    wb1.cell(i+2,2,school_2014_2015[i])
    wb1.cell(i + 2, 3, people_name_2014_2015[i])
    wb1.cell(i + 2, 4, department_2014_2015[i])
    wb1.cell(i + 2, 5, project_2014_2015[i])
    wb1.cell(i + 2, 6, code_2014_2015[i])
    wb1.cell(i + 2, 7, gender_2014_2015[i])
    wb1.cell(i + 2, 8, nation_2014_2015[i])
    wb1.cell(i + 2, 9, date_2014_2015[i])
    wb1.cell(i + 2, 10, book_2014_2015[i])
    wb1.cell(i + 2, 11, area_2014_2015[i])
    wb1.cell(i + 2, 12, '2014-2015')
    wb1.cell(i + 2, 13, class_2014_2015[i])
wb.save("2014—2015学年度本专科生国家奖学金.xlsx")
'''

'''
#2015-2016
path_2015_2016_read_country='F:\奈园代码\本专科生国家奖学金\原始数据\p2015—2016学年度本专科生国家奖学金\中央高校'
path_2015_2016_read_city='F:\奈园代码\本专科生国家奖学金\原始数据\p2015—2016学年度本专科生国家奖学金\地方高校'
path_2015_2016_write='F:\奈园代码\本专科生国家奖学金\处理数据'
filename_2015_2016_country=os.listdir(path_2015_2016_read_country)
filename_2015_2016_city=os.listdir(path_2015_2016_read_city)


people_name_2015_2016=[]
school_2015_2016=[]
department_2015_2016=[]
project_2015_2016=[]
code_2015_2016=[]
gender_2015_2016=[]
nation_2015_2016=[]
date_2015_2016=[]
book_2015_2016=[]
#地区/中央
class_2015_2016=[]
class_2015_2016_country_index=[]
class_2015_2016_city_index=[]
class_2015_2016_country_index_name=[]
class_2015_2016_city_index_name=[]
for i in range(len(filename_2015_2016_country)):
    os.chdir(path_2015_2016_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2015_2016_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    for j in range(2,row_count):
        class_2015_2016_country_index.append(sheet.cell(j,1).value)
for i in range(len(class_2015_2016_country_index)):
    class_2015_2016_country_index_name.append('中央高校')
for i in range(len(filename_2015_2016_city)):
    os.chdir(path_2015_2016_read_city)
    workbook = xlrd.open_workbook(r'%s' % filename_2015_2016_city[i])
    sheet = workbook.sheet_by_index(0)
    row_count = sheet.nrows
    col_count = sheet.ncols
    for j in range(2, row_count):
        class_2015_2016_city_index.append(sheet.cell(j, 1).value)
for i in range(len(class_2015_2016_city_index)):
    class_2015_2016_city_index_name.append('地方高校')
class_2015_2016=class_2015_2016_country_index_name+class_2015_2016_city_index_name
#加地区
area_2015_2016=[]
area_2015_2016_index_country=[]
area_2015_2016_index_city=[]
area_2015_2016_index_country_name=[]
area_2015_2016_index_city_name=[]
for i in range(len(filename_2015_2016_country)):
    os.chdir(path_2015_2016_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2015_2016_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    area_2015_2016_index_country.append(row_count-2)
for i in range(len(area_2015_2016_index_country)):
    for j in range(area_2015_2016_index_country[i]):
        area_2015_2016_index_country_name.append('')
for i in range(len(filename_2015_2016_city)):
    os.chdir(path_2015_2016_read_city)
    workbook=xlrd.open_workbook(r'%s'%filename_2015_2016_city[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    area_2015_2016_index_city.append(row_count-2)
for i in range(len(area_2015_2016_index_city)):
    for j in range(area_2015_2016_index_city[i]):
        area_2015_2016_index_city_name.append(filename_2015_2016_city[i])
area_2015_2016=area_2015_2016_index_country_name+area_2015_2016_index_city_name


#读全部文件
for i in range(len(filename_2015_2016_country)):
    os.chdir(path_2015_2016_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2015_2016_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    for j in range(2,row_count):
        people_name_2015_2016.append(sheet.cell(j,1).value)
        school_2015_2016.append(sheet.cell(j, 2).value)
        department_2015_2016.append(sheet.cell(j, 3).value)
        project_2015_2016.append(sheet.cell(j, 4).value)
        code_2015_2016.append(sheet.cell(j, 5).value)
        gender_2015_2016.append(sheet.cell(j, 6).value)
        nation_2015_2016.append(sheet.cell(j, 7).value)
        date_2015_2016.append(sheet.cell(j, 8).value)
        book_2015_2016.append(sheet.cell(j, 9).value)
for i in range(len(filename_2015_2016_city)):
    os.chdir(path_2015_2016_read_city)
    workbook=xlrd.open_workbook(r'%s'%filename_2015_2016_city[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    print(row_count,col_count)
    for j in range(2,row_count):
        people_name_2015_2016.append(sheet.cell(j,1).value)
        school_2015_2016.append(sheet.cell(j, 2).value)
        department_2015_2016.append(sheet.cell(j, 3).value)
        project_2015_2016.append(sheet.cell(j, 4).value)
        code_2015_2016.append(sheet.cell(j, 5).value)
        gender_2015_2016.append(sheet.cell(j, 6).value)
        nation_2015_2016.append(sheet.cell(j, 7).value)
        date_2015_2016.append(sheet.cell(j, 8).value)
        book_2015_2016.append(sheet.cell(j, 9).value)

#写

os.chdir(path_2015_2016_write)
wb=load_workbook("2015—2016学年度本专科生国家奖学金.xlsx")
wb1=wb.active
for i in range(len(people_name_2015_2016)):
    wb1.cell(i+2,1,i+1)
    wb1.cell(i+2,2,school_2015_2016[i])
    wb1.cell(i + 2, 3, people_name_2015_2016[i])
    wb1.cell(i + 2, 4, department_2015_2016[i])
    wb1.cell(i + 2, 5, project_2015_2016[i])
    wb1.cell(i + 2, 6, code_2015_2016[i])
    wb1.cell(i + 2, 7, gender_2015_2016[i])
    wb1.cell(i + 2, 8, nation_2015_2016[i])
    wb1.cell(i + 2, 9, date_2015_2016[i])
    wb1.cell(i + 2, 10, book_2015_2016[i])
    wb1.cell(i + 2, 11, area_2015_2016[i])
    wb1.cell(i + 2, 12, '2014-2015')
    wb1.cell(i + 2, 13, class_2015_2016[i])
wb.save("2015—2016学年度本专科生国家奖学金.xlsx")
'''

'''
#2016-2017
path_2016_2017_read_country='F:\奈园代码\本专科生国家奖学金\原始数据\p2016—2017学年度本专科生国家奖学金\p2017中央高校-19993人'
path_2016_2017_read_city='F:\奈园代码\本专科生国家奖学金\原始数据\p2016—2017学年度本专科生国家奖学金\p2017地方高校 -29997人'
path_2016_2017_write='F:\奈园代码\本专科生国家奖学金\处理数据'
filename_2016_2017_country=os.listdir(path_2016_2017_read_country)
filename_2016_2017_city=os.listdir(path_2016_2017_read_city)

people_name_2016_2017=[]
school_2016_2017=[]
department_2016_2017=[]
project_2016_2017=[]
code_2016_2017=[]
gender_2016_2017=[]
nation_2016_2017=[]
date_2016_2017=[]
book_2016_2017=[]
#地区/中央
class_2016_2017=[]
class_2016_2017_country_index=[]
class_2016_2017_city_index=[]
class_2016_2017_country_index_name=[]
class_2016_2017_city_index_name=[]
for i in range(len(filename_2016_2017_country)):
    os.chdir(path_2016_2017_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2016_2017_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    print(row_count,col_count)
    for j in range(2,row_count):
        class_2016_2017_country_index.append(sheet.cell(j,1).value)
for i in range(len(class_2016_2017_country_index)):
    class_2016_2017_country_index_name.append('中央高校')
for i in range(len(filename_2016_2017_city)):
    os.chdir(path_2016_2017_read_city)
    workbook = xlrd.open_workbook(r'%s' % filename_2016_2017_city[i])
    sheet = workbook.sheet_by_index(0)
    row_count = sheet.nrows
    col_count = sheet.ncols
    print(row_count, col_count)
    for j in range(2, row_count):
        class_2016_2017_city_index.append(sheet.cell(j, 1).value)
for i in range(len(class_2016_2017_city_index)):
    class_2016_2017_city_index_name.append('地方高校')
class_2016_2017=class_2016_2017_country_index_name+class_2016_2017_city_index_name
#加地区
area_2016_2017=[]
area_2016_2017_index_country=[]
area_2016_2017_index_city=[]
area_2016_2017_index_country_name=[]
area_2016_2017_index_city_name=[]
for i in range(len(filename_2016_2017_country)):
    os.chdir(path_2016_2017_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2016_2017_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    area_2016_2017_index_country.append(row_count-2)
for i in range(len(area_2016_2017_index_country)):
    for j in range(area_2016_2017_index_country[i]):
        area_2016_2017_index_country_name.append('')
for i in range(len(filename_2016_2017_city)):
    os.chdir(path_2016_2017_read_city)
    workbook=xlrd.open_workbook(r'%s'%filename_2016_2017_city[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    area_2016_2017_index_city.append(row_count-2)
for i in range(len(area_2016_2017_index_city)):
    for j in range(area_2016_2017_index_city[i]):
        area_2016_2017_index_city_name.append(filename_2016_2017_city[i])
area_2016_2017=area_2016_2017_index_country_name+area_2016_2017_index_city_name


#读全部文件
for i in range(len(filename_2016_2017_country)):
    os.chdir(path_2016_2017_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2016_2017_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    for j in range(2,row_count):
        people_name_2016_2017.append(sheet.cell(j,1).value)
        school_2016_2017.append(sheet.cell(j, 2).value)
        department_2016_2017.append(sheet.cell(j, 3).value)
        project_2016_2017.append(sheet.cell(j, 4).value)
        code_2016_2017.append(sheet.cell(j, 5).value)
        gender_2016_2017.append(sheet.cell(j, 6).value)
        nation_2016_2017.append(sheet.cell(j, 7).value)
        date_2016_2017.append(sheet.cell(j, 8).value)
        book_2016_2017.append(sheet.cell(j, 9).value)
for i in range(len(filename_2016_2017_city)):
    os.chdir(path_2016_2017_read_city)
    workbook=xlrd.open_workbook(r'%s'%filename_2016_2017_city[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    for j in range(2,row_count):
        people_name_2016_2017.append(sheet.cell(j,1).value)
        school_2016_2017.append(sheet.cell(j, 2).value)
        department_2016_2017.append(sheet.cell(j, 3).value)
        project_2016_2017.append(sheet.cell(j, 4).value)
        code_2016_2017.append(sheet.cell(j, 5).value)
        gender_2016_2017.append(sheet.cell(j, 6).value)
        nation_2016_2017.append(sheet.cell(j, 7).value)
        date_2016_2017.append(sheet.cell(j, 8).value)
        book_2016_2017.append(sheet.cell(j, 9).value)

#写

os.chdir(path_2016_2017_write)
wb=load_workbook("2016—2017学年度本专科生国家奖学金.xlsx")
wb1=wb.active
for i in range(len(people_name_2016_2017)):
    wb1.cell(i+2,1,i+1)
    wb1.cell(i+2,2,school_2016_2017[i])
    wb1.cell(i + 2, 3, people_name_2016_2017[i])
    wb1.cell(i + 2, 4, department_2016_2017[i])
    wb1.cell(i + 2, 5, project_2016_2017[i])
    wb1.cell(i + 2, 6, code_2016_2017[i])
    wb1.cell(i + 2, 7, gender_2016_2017[i])
    wb1.cell(i + 2, 8, nation_2016_2017[i])
    wb1.cell(i + 2, 9, date_2016_2017[i])
    wb1.cell(i + 2, 10, book_2016_2017[i])
    wb1.cell(i + 2, 11, area_2016_2017[i])
    wb1.cell(i + 2, 12, '2014-2015')
    wb1.cell(i + 2, 13, class_2016_2017[i])
wb.save("2016—2017学年度本专科生国家奖学金.xlsx")
'''

'''
#2017-2018
path_2017_2018_read_country='F:\奈园代码\本专科生国家奖学金\原始数据\p2017—2018学年度本专科生国家奖学金\p1中央高校'
path_2017_2018_read_city='F:\奈园代码\本专科生国家奖学金\原始数据\p2017—2018学年度本专科生国家奖学金\p2地方高校'
path_2017_2018_write='F:\奈园代码\本专科生国家奖学金\处理数据'
filename_2017_2018_country=os.listdir(path_2017_2018_read_country)
filename_2017_2018_city=os.listdir(path_2017_2018_read_city)

people_name_2017_2018=[]
school_2017_2018=[]
department_2017_2018=[]
project_2017_2018=[]
code_2017_2018=[]
gender_2017_2018=[]
nation_2017_2018=[]
date_2017_2018=[]
book_2017_2018=[]
#地区/中央
class_2017_2018=[]
class_2017_2018_country_index=[]
class_2017_2018_city_index=[]
class_2017_2018_country_index_name=[]
class_2017_2018_city_index_name=[]
for i in range(len(filename_2017_2018_country)):
    os.chdir(path_2017_2018_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2017_2018_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    print(filename_2017_2018_country[i],row_count,col_count)
    for j in range(2,row_count):
        class_2017_2018_country_index.append(sheet.cell(j,1).value)
for i in range(len(class_2017_2018_country_index)):
    class_2017_2018_country_index_name.append('中央高校')
for i in range(len(filename_2017_2018_city)):
    os.chdir(path_2017_2018_read_city)
    workbook = xlrd.open_workbook(r'%s' % filename_2017_2018_city[i])
    sheet = workbook.sheet_by_index(0)
    row_count = sheet.nrows
    col_count = sheet.ncols
    print(filename_2017_2018_city[i],row_count, col_count)
    for j in range(2, row_count):
        class_2017_2018_city_index.append(sheet.cell(j, 1).value)
for i in range(len(class_2017_2018_city_index)):
    class_2017_2018_city_index_name.append('地方高校')
class_2017_2018=class_2017_2018_country_index_name+class_2017_2018_city_index_name
#加地区
area_2017_2018=[]
area_2017_2018_index_country=[]
area_2017_2018_index_city=[]
area_2017_2018_index_country_name=[]
area_2017_2018_index_city_name=[]
for i in range(len(filename_2017_2018_country)):
    os.chdir(path_2017_2018_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2017_2018_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    area_2017_2018_index_country.append(row_count-2)
for i in range(len(area_2017_2018_index_country)):
    for j in range(area_2017_2018_index_country[i]):
        area_2017_2018_index_country_name.append('')
for i in range(len(filename_2017_2018_city)):
    os.chdir(path_2017_2018_read_city)
    workbook=xlrd.open_workbook(r'%s'%filename_2017_2018_city[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    area_2017_2018_index_city.append(row_count-2)
for i in range(len(area_2017_2018_index_city)):
    for j in range(area_2017_2018_index_city[i]):
        area_2017_2018_index_city_name.append(filename_2017_2018_city[i])
area_2017_2018=area_2017_2018_index_country_name+area_2017_2018_index_city_name


#读全部文件
for i in range(len(filename_2017_2018_country)):
    os.chdir(path_2017_2018_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2017_2018_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    for k in range(row_count-2):
        school_2017_2018.append(filename_2017_2018_country[i])
    for j in range(2,row_count):
        people_name_2017_2018.append(sheet.cell(j,1).value)
        # school_2017_2018.append(sheet.cell(j, 2).value)
        department_2017_2018.append(sheet.cell(j, 2).value)
        project_2017_2018.append(sheet.cell(j, 3).value)
        code_2017_2018.append(sheet.cell(j, 4).value)
        gender_2017_2018.append(sheet.cell(j, 5).value)
        nation_2017_2018.append(sheet.cell(j, 6).value)
        date_2017_2018.append(sheet.cell(j, 7).value)
        book_2017_2018.append(sheet.cell(j, 8).value)
print(school_2017_2018)
for i in range(len(filename_2017_2018_city)):
    os.chdir(path_2017_2018_read_city)
    workbook=xlrd.open_workbook(r'%s'%filename_2017_2018_city[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    for j in range(2,row_count):
        people_name_2017_2018.append(sheet.cell(j,1).value)
        school_2017_2018.append(sheet.cell(j, 2).value)
        department_2017_2018.append(sheet.cell(j, 3).value)
        project_2017_2018.append(sheet.cell(j, 4).value)
        code_2017_2018.append(sheet.cell(j, 5).value)
        gender_2017_2018.append(sheet.cell(j, 6).value)
        nation_2017_2018.append(sheet.cell(j, 7).value)
        date_2017_2018.append(sheet.cell(j, 8).value)
        book_2017_2018.append(sheet.cell(j, 9).value)

#写
os.chdir(path_2017_2018_write)
wb=load_workbook("2017—2018学年度本专科生国家奖学金.xlsx")
wb1=wb.active
for i in range(len(people_name_2017_2018)):
    wb1.cell(i+2,1,i+1)
    wb1.cell(i+2,2,school_2017_2018[i])
    wb1.cell(i + 2, 3, people_name_2017_2018[i])
    wb1.cell(i + 2, 4, department_2017_2018[i])
    wb1.cell(i + 2, 5, project_2017_2018[i])
    wb1.cell(i + 2, 6, code_2017_2018[i])
    wb1.cell(i + 2, 7, gender_2017_2018[i])
    wb1.cell(i + 2, 8, nation_2017_2018[i])
    wb1.cell(i + 2, 9, date_2017_2018[i])
    wb1.cell(i + 2, 10, book_2017_2018[i])
    wb1.cell(i + 2, 11, area_2017_2018[i])
    wb1.cell(i + 2, 12, '2017-2018')
    wb1.cell(i + 2, 13, class_2017_2018[i])
wb.save("2017—2018学年度本专科生国家奖学金.xlsx")
'''

# '''
#2018-2019
path_2018_2019_read_country='F:\奈园代码\本专科生国家奖学金\原始数据\p2018—2019学年度本专科生国家奖学金\中央高校'
path_2018_2019_read_city='F:\奈园代码\本专科生国家奖学金\原始数据\p2018—2019学年度本专科生国家奖学金\地方高校'
path_2018_2019_write='F:\奈园代码\本专科生国家奖学金\处理数据'
filename_2018_2019_country=os.listdir(path_2018_2019_read_country)
filename_2018_2019_city=os.listdir(path_2018_2019_read_city)

people_name_2018_2019=[]
school_2018_2019=[]
department_2018_2019=[]
project_2018_2019=[]
code_2018_2019=[]
gender_2018_2019=[]
nation_2018_2019=[]
date_2018_2019=[]
book_2018_2019=[]
#地区/中央
class_2018_2019=[]
class_2018_2019_country_index=[]
class_2018_2019_city_index=[]
class_2018_2019_country_index_name=[]
class_2018_2019_city_index_name=[]
for i in range(len(filename_2018_2019_country)):
    os.chdir(path_2018_2019_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2018_2019_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    print(filename_2018_2019_country[i],row_count,col_count)
    for j in range(2,row_count):
        class_2018_2019_country_index.append(sheet.cell(j,1).value)
for i in range(len(class_2018_2019_country_index)):
    class_2018_2019_country_index_name.append('中央高校')
for i in range(len(filename_2018_2019_city)):
    os.chdir(path_2018_2019_read_city)
    workbook = xlrd.open_workbook(r'%s' % filename_2018_2019_city[i])
    sheet = workbook.sheet_by_index(0)
    row_count = sheet.nrows
    col_count = sheet.ncols
    print(filename_2018_2019_city[i],row_count, col_count)
    for j in range(2, row_count):
        class_2018_2019_city_index.append(sheet.cell(j, 1).value)
for i in range(len(class_2018_2019_city_index)):
    class_2018_2019_city_index_name.append('地方高校')
class_2018_2019=class_2018_2019_country_index_name+class_2018_2019_city_index_name
#加地区
area_2018_2019=[]
area_2018_2019_index_country=[]
area_2018_2019_index_city=[]
area_2018_2019_index_country_name=[]
area_2018_2019_index_city_name=[]
for i in range(len(filename_2018_2019_country)):
    os.chdir(path_2018_2019_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2018_2019_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    area_2018_2019_index_country.append(row_count-2)
for i in range(len(area_2018_2019_index_country)):
    for j in range(area_2018_2019_index_country[i]):
        area_2018_2019_index_country_name.append('')
for i in range(len(filename_2018_2019_city)):
    os.chdir(path_2018_2019_read_city)
    workbook=xlrd.open_workbook(r'%s'%filename_2018_2019_city[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    area_2018_2019_index_city.append(row_count-2)
for i in range(len(area_2018_2019_index_city)):
    for j in range(area_2018_2019_index_city[i]):
        area_2018_2019_index_city_name.append(filename_2018_2019_city[i])
area_2018_2019=area_2018_2019_index_country_name+area_2018_2019_index_city_name


#读全部文件
for i in range(len(filename_2018_2019_country)):
    os.chdir(path_2018_2019_read_country)
    workbook=xlrd.open_workbook(r'%s'%filename_2018_2019_country[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    for j in range(2,row_count):
        people_name_2018_2019.append(sheet.cell(j,1).value)
        school_2018_2019.append(sheet.cell(j, 2).value)
        department_2018_2019.append(sheet.cell(j, 3).value)
        project_2018_2019.append(sheet.cell(j, 4).value)
        code_2018_2019.append(sheet.cell(j, 5).value)
        gender_2018_2019.append(sheet.cell(j, 6).value)
        nation_2018_2019.append(sheet.cell(j, 7).value)
        date_2018_2019.append(sheet.cell(j, 8).value)
        book_2018_2019.append(sheet.cell(j, 9).value)
for i in range(len(filename_2018_2019_city)):
    os.chdir(path_2018_2019_read_city)
    workbook=xlrd.open_workbook(r'%s'%filename_2018_2019_city[i])
    sheet=workbook.sheet_by_index(0)
    row_count=sheet.nrows
    col_count=sheet.ncols
    for j in range(2,row_count):
        people_name_2018_2019.append(sheet.cell(j,1).value)
        school_2018_2019.append(sheet.cell(j, 2).value)
        department_2018_2019.append(sheet.cell(j, 3).value)
        project_2018_2019.append(sheet.cell(j, 4).value)
        code_2018_2019.append(sheet.cell(j, 5).value)
        gender_2018_2019.append(sheet.cell(j, 6).value)
        nation_2018_2019.append(sheet.cell(j, 7).value)
        date_2018_2019.append(sheet.cell(j, 8).value)
        book_2018_2019.append(sheet.cell(j, 9).value)

#写

os.chdir(path_2018_2019_write)
wb=load_workbook("2018—2019学年度本专科生国家奖学金.xlsx")
wb1=wb.active
for i in range(len(people_name_2018_2019)):
    wb1.cell(i+2,1,i+1)
    wb1.cell(i+2,2,school_2018_2019[i])
    wb1.cell(i + 2, 3, people_name_2018_2019[i])
    wb1.cell(i + 2, 4, department_2018_2019[i])
    wb1.cell(i + 2, 5, project_2018_2019[i])
    wb1.cell(i + 2, 6, code_2018_2019[i])
    wb1.cell(i + 2, 7, gender_2018_2019[i])
    wb1.cell(i + 2, 8, nation_2018_2019[i])
    wb1.cell(i + 2, 9, date_2018_2019[i])
    wb1.cell(i + 2, 10, book_2018_2019[i])
    wb1.cell(i + 2, 11, area_2018_2019[i])
    wb1.cell(i + 2, 12, '2018-2019')
    wb1.cell(i + 2, 13, class_2018_2019[i])
wb.save("2018—2019学年度本专科生国家奖学金.xlsx")


path_2018_2019_write='F:\奈园代码\本专科生国家奖学金\处理数据'
os.chdir(path_2018_2019_write)
workbook=xlrd.open_workbook(r'2018—2019学年度本专科生国家奖学金.xlsx')
sheet=workbook.sheet_by_index(0)
row_count=sheet.nrows
col_count=sheet.ncols
date=[]
import datetime
for j in range(1,row_count):
    date.append(sheet.cell(j,8).value)
# print(date)
def date_transform(dates):
    delta=datetime.timedelta(days=dates)
    today=datetime.datetime.strptime('1900-1','%Y-%m')+delta
    return datetime.datetime.strftime(today,'%Y-%m')
number=[x for x in range(50000)]
for i in range(len(date)):
    for j in number:
        if date[i]==j:
            date[i]=date_transform(date[i])
# print(date)
wb=load_workbook("2018—2019学年度本专科生国家奖学金.xlsx")
wb1=wb.active
for i in range(len(date)):
    wb1.cell(i+2,14,date[i])
wb.save("2018—2019学年度本专科生国家奖学金.xlsx")
# '''