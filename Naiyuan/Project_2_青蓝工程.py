import os
import pandas as pd
import docx
from docx import Document
import xlrd
from xlutils.copy import copy
import openpyxl
from openpyxl import load_workbook

Result_1=pd.read_excel('江苏省“青蓝工程”优秀教学团队.xlsx')
Result_2=pd.read_excel('江苏省“青蓝工程”优秀青年骨干教师.xlsx')
Result_3=pd.read_excel('江苏省“青蓝工程”中青年学术带头人.xlsx')

# data_2019=pd.read_excel('2019_work.xlsx')
# data_2018=pd.read_excel('...')
# data_2017=pd.read_excel('...')

##########################################2016
gugan_teacher_2016=[]
gugan_teacher_school_2016=[]

xueke_leader_2016=[]
xueke_leader_school_2016=[]

wb1=xlrd.open_workbook('2016_work.xlsx')
sheet=wb1.sheet_by_name('Sheet1')

x=[]
for a in range(266):
    cells = sheet.row_values(a)
    data = str(cells[0])
    x.append(data)
# print(len(x))#266
single=[]
two=[]
for i in range(133):
    single.append(i*2)
    two.append(i*2+1)
for i in range(len(single)):
    gugan_teacher_school_2016.append(x[single[i]])
for i in range(len(two)):
    gugan_teacher_2016.append(x[two[i]])
# print(gugan_teacher_2016)
# print(gugan_teacher_school_2016)

x=[]
for a in range(266,458):
    cells = sheet.row_values(a)
    data = str(cells[0])
    x.append(data)
# print(len(x))#192
single=[]
two=[]
for i in range(96):
    single.append(i*2)
    two.append(i*2+1)
for i in range(len(single)):
    xueke_leader_school_2016.append(x[single[i]])
for i in range(len(two)):
    xueke_leader_2016.append(x[two[i]])
print(xueke_leader_school_2016)
print(xueke_leader_2016)



# ##########################################2017
# gugan_teacher_2017=[]
# gugan_teacher_school_2017=[]
#
# xueke_leader_2017=[]
# xueke_leader_school_2017=[]
#
# teach_team_name_2017=[]
# teach_team_teacher_2017=[]
# teach_team_school_2017=[]
#
# wb1=xlrd.open_workbook('2017_work.xlsx')
# sheet=wb1.sheet_by_name('Sheet1')
#
# x=[]
# for a in range(304):
#     cells = sheet.row_values(a)
#     data = str(cells[0])
#     x.append(data)
# # print(len(x))#304
# single=[]
# two=[]
# for i in range(152):
#     single.append(i*2)
#     two.append(i*2+1)
# for i in range(len(single)):
#     gugan_teacher_school_2017.append(x[single[i]])
# for i in range(len(two)):
#     gugan_teacher_2017.append(x[two[i]])
# print(gugan_teacher_2017)
# print(gugan_teacher_school_2017)
#
# x=[]
# for a in range(304,504):
#     cells = sheet.row_values(a)
#     data = str(cells[0])
#     x.append(data)
# # print(len(x))#200
# single=[]
# two=[]
# for i in range(100):
#     single.append(i*2)
#     two.append(i*2+1)
# for i in range(len(single)):
#     xueke_leader_school_2017.append(x[single[i]])
# for i in range(len(two)):
#     xueke_leader_2017.append(x[two[i]])
# print(xueke_leader_school_2017)
# print(xueke_leader_2017)
#
# for a in range(504,564):
#     cells=sheet.row_values(a)
#     data=str(cells[0])
#     teach_team_school_2017.append(data)
#     data=str(cells[1])
#     teach_team_teacher_2017.append(data)
#     data=str(cells[2])
#     teach_team_name_2017.append(data)
# print(teach_team_teacher_2017,teach_team_name_2017,teach_team_school_2017)


# ###############################2018
# gugan_teacher_2018=[]
# gugan_teacher_school_2018=[]
#
# xueke_leader_2018=[]
# xueke_leader_school_2018=[]
#
# teach_team_name_2018=[]
# teach_team_teacher_2018=[]
# teach_team_school_2018=[]
#
# wb1=xlrd.open_workbook('2018_work.xlsx')
# sheet=wb1.sheet_by_name('Sheet1')
#
# x=[]
# for a in range(314):
#     cells = sheet.row_values(a)
#     data = str(cells[0])
#     x.append(data)
# # print(len(x))#314
# single=[]
# two=[]
# for i in range(157):
#     single.append(i*2)
#     two.append(i*2+1)
# for i in range(len(single)):
#     gugan_teacher_school_2018.append(x[single[i]])
# for i in range(len(two)):
#     gugan_teacher_2018.append(x[two[i]])
# print(gugan_teacher_2018)
# print(gugan_teacher_school_2018)
#
# x=[]
# for a in range(314,502):
#     cells = sheet.row_values(a)
#     data = str(cells[0])
#     x.append(data)
# # print(len(x))#188
# single=[]
# two=[]
# for i in range(94):
#     single.append(i*2)
#     two.append(i*2+1)
# for i in range(len(single)):
#     xueke_leader_school_2018.append(x[single[i]])
# for i in range(len(two)):
#     xueke_leader_2018.append(x[two[i]])
# print(xueke_leader_school_2018)
# print(xueke_leader_2018)
#
# for a in range(502,569):
#     cells=sheet.row_values(a)
#     data=str(cells[0])
#     teach_team_school_2018.append(data)
#     data=str(cells[1])
#     teach_team_teacher_2018.append(data)
#     data=str(cells[2])
#     teach_team_name_2018.append(data)
# print(teach_team_teacher_2018,teach_team_name_2018,teach_team_school_2018)


# ###########################################2019
# gugan_teacher_2019=[]
# gugan_teacher_school_2019=[]
#
# xueke_leader_2019=[]
# xueke_leader_school_2019=[]
#
# teach_team_name_2019=[]
# teach_team_teacher_2019=[]
# teach_team_school_2019=[]
#
# wb1=xlrd.open_workbook('2019_work.xlsx')
# sheet=wb1.sheet_by_name('Sheet1')
#
# x=[]
# for a in range(306):
#     cells = sheet.row_values(a)
#     data = str(cells[0])
#     x.append(data)
# # print(len(x))#306
# single=[]
# two=[]
# for i in range(153):
#     single.append(i*2)
#     two.append(i*2+1)
# for i in range(len(single)):
#     gugan_teacher_school_2019.append(x[single[i]])
# for i in range(len(two)):
#     gugan_teacher_2019.append(x[two[i]])
# print(gugan_teacher_2019)
# print(gugan_teacher_school_2019)
#
# x=[]
# for a in range(306,488):
#     cells = sheet.row_values(a)
#     data = str(cells[0])
#     x.append(data)
# # print(len(x))#91
# single=[]
# two=[]
# for i in range(91):
#     single.append(i*2)
#     two.append(i*2+1)
# for i in range(len(single)):
#     xueke_leader_school_2019.append(x[single[i]])
# for i in range(len(two)):
#     xueke_leader_2019.append(x[two[i]])
# print(xueke_leader_school_2019)
# print(xueke_leader_2019)
#
# for a in range(488,558):
#     cells=sheet.row_values(a)
#     data=str(cells[0])
#     teach_team_school_2019.append(data)
#     data=str(cells[1])
#     teach_team_teacher_2019.append(data)
#     data=str(cells[2])
#     teach_team_name_2019.append(data)
# print(teach_team_teacher_2019,teach_team_name_2019,teach_team_school_2019)



wb2=load_workbook('2016_骨干.xlsx')
wb3=wb2.active
for i in range(len(gugan_teacher_2016)):
    wb3.cell(i+2,2,gugan_teacher_school_2016[i])
    wb3.cell(i+2,3,gugan_teacher_2016[i])
    wb3.cell(i+2,4,'2016')
wb2.save('2016_骨干.xlsx')

wb6=load_workbook('2016_leader.xlsx')
wb7=wb6.active
for i in range(len(xueke_leader_school_2016)):
    wb7.cell(i+2,2,xueke_leader_school_2016[i])
    wb7.cell(i+2,3,xueke_leader_2016[i])
    wb7.cell(i+2,4,'2016')
wb6.save('2016_leader.xlsx')

'''
wb2=load_workbook('江苏省“青蓝工程”优秀青年骨干教师.xlsx')
wb3=wb2.active
for i in range(len(gugan_teacher_2017)):
    wb3.cell(i+2,2,gugan_teacher_school_2017[i])
    wb3.cell(i+2,3,gugan_teacher_2017[i])
    wb3.cell(i+2,4,'2017')
for i in range(len(gugan_teacher_2018)):
    wb3.cell(i + 154, 2, gugan_teacher_school_2018[i])
    wb3.cell(i + 154, 3, gugan_teacher_2018[i])
    wb3.cell(i + 154, 4, '2018')
for i in range(len(gugan_teacher_2019)):
    wb3.cell(i + 311, 2, gugan_teacher_school_2018[i])
    wb3.cell(i + 311, 3, gugan_teacher_2019[i])
    wb3.cell(i + 311, 4, '2019')
wb2.save('江苏省“青蓝工程”优秀青年骨干教师.xlsx')

wb4=load_workbook('江苏省“青蓝工程”优秀教学团队.xlsx')
wb5=wb4.active
for i in range(len(teach_team_school_2017)):
    wb5.cell(i+2,2,teach_team_school_2017[i])
    wb5.cell(i+2,3,teach_team_teacher_2017[i])
    wb5.cell(i+2,4,'2017')
    wb5.cell(i+2,5,teach_team_name_2017[i])
for i in range(len(teach_team_school_2018)):
    wb5.cell(i + 62, 2, teach_team_school_2018[i])
    wb5.cell(i + 62, 3, teach_team_teacher_2018[i])
    wb5.cell(i + 62, 4, '2018')
    wb5.cell(i + 62, 5, teach_team_name_2018[i])
for i in range(len(teach_team_school_2019)):
    wb5.cell(i + 129, 2, teach_team_school_2019[i])
    wb5.cell(i + 129, 3, teach_team_teacher_2019[i])
    wb5.cell(i + 129, 4, '2019')
    wb5.cell(i + 129, 5, teach_team_name_2019[i])
wb4.save('江苏省“青蓝工程”优秀教学团队.xlsx')

wb6=load_workbook('江苏省“青蓝工程”中青年学术带头人.xlsx')
wb7=wb6.active
for i in range(len(xueke_leader_school_2017)):
    wb7.cell(i+2,2,xueke_leader_school_2017[i])
    wb7.cell(i+2,3,xueke_leader_2017[i])
    wb7.cell(i+2,4,'2017')
for i in range(len(xueke_leader_school_2018)):
    wb7.cell(i + 102, 2, xueke_leader_school_2018[i])
    wb7.cell(i + 102, 3, xueke_leader_2018[i])
    wb7.cell(i + 102, 4, '2018')
for i in range(len(xueke_leader_school_2019)):
    wb7.cell(i + 196, 2, xueke_leader_school_2019[i])
    wb7.cell(i + 196, 3, xueke_leader_2019[i])
    wb7.cell(i + 196, 4, '2019')
wb6.save('江苏省“青蓝工程”中青年学术带头人.xlsx')
'''