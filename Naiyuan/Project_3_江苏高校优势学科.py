import os
import pandas as pd
import docx
import xlrd
from docx import Document
from xlutils.copy import copy
import openpyxl
from openpyxl import load_workbook
import re

wb=xlrd.open_workbook('333.xlsx')
sheet=wb.sheet_by_name('2018')
name=[]
school=[]
for i in range(sheet.nrows):
    cells=sheet.row_values(i)
    data=str(cells[1])
    data2=str(cells[0])
    name.append(data)
    school.append(data2)
def cut_text(text,lenth):
    textArr=re.findall('.{'+str(lenth)+'}',text)
    textArr.append(text[(len(textArr)*lenth):])
    return textArr
name_new=[]
name_new1=[]
name_n=[]
for j in range(len(name)):
    name_new.append(cut_text(name[j],4))

for k in range(len(name_new)):
    for j in range(len(name_new[k])):
        # print(name_new[k][j])
        name_new1.append(name_new[k][j])
name_new1=[i for i in name_new1 if i != '']


# wb2=load_workbook('1.xlsx')
# wb3=wb2.active
# for i in range(len(name_new1)):
#     wb3.cell(i+1,1,str(name_new1[i]))
# wb2.save("1.xlsx")

wb3=load_workbook('2.xlsx')
wb4=wb3.active

a=[]
for j in name_new:
    k=len(j)
    a.append(k)
# print(name_new)
with open('1.txt','w') as f:
    f.write(str(name_new))
# print(a)
count=0
for i in range(len(a)):
    count+=a[i]
print(count)

# print(a)#单个学校的个数
# # print(school)
# print(name_new)
# print(len(name_new1))
# # print(name_new1)
# # print(len(name_new1))
#
#写一个列表，包含3703个学校
new=[]
for i in range(len(school)):
    new.append([school[i]])
new_2=[]
for i in range(len(school)):
    new_2.append(new[i]*a[i])
# print(new_2)
new_3=[]
for k in range(len(school)):
    for j in range(len(new_2[k])):
       new_3.append(new_2[k][j])
print(new_3)

for i in range(len(new_3)):
    wb4.cell(i+1,1,new_3[i])

wb3.save('2.xlsx')