import requests
import re
import lxml
import os
from lxml import etree
from openpyxl import load_workbook
import xlrd

path='F:\奈园代码'
os.chdir(path)
workbook=xlrd.open_workbook(r'科教文章库爬取链接地址.xlsx')
sheet=workbook.sheet_by_index(0)
row_count=sheet.nrows
col_count=sheet.ncols
single=[]
double=[]
college=[]
website=[]
for j in range(0,row_count):
    if j%2==0:
        single.append(j)
    else:
        double.append(j)
        # single.append(sheet.cell(j,0).value)
# print(single,double)
for i in single:
    college.append(sheet.cell(i,0).value)
for i in double:
    website.append(sheet.cell(i,0).value)
print(college,website)

wb=load_workbook('1.xlsx')
wb1=wb.active
for i in range(len(college)):
    wb1.cell(i+1,1,college[i])
    wb1.cell(i+1,2,website[i])
wb.save('1.xlsx')